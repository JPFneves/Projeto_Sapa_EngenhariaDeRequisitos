"""
SAPA v7.0 — Sistema de Automação de Presença Acadêmica
UNISEPE — ADS 3° Semestre

v7.0:
  ✅ Leitor USB universal (qualquer leitor que emule teclado)
  ✅ Painel do leitor com detecção real de dispositivos HID conectados
  ✅ Validação de RA: só aceita padrão UNISEPE (ano + RA numérico) — outros códigos ignorados
  ✅ Registro único por aluno: Entrada + Saída na mesma linha
  ✅ Painel professor: todos os alunos listados, CSV por dia ou histórico completo para Power BI
  ✅ Seletor de pasta nativo ao baixar CSV
  ✅ Fechar o X pergunta se quer enviar relatório ao professor da aula
  ✅ Login admin persistente na sessão
  ✅ Calendário visual próprio sem dependência externa
  ✅ Indicador Wi-Fi online/offline em tempo real
  ✅ Tela cheia F11, clock em PT-BR, engrenagem de configurações
"""

import sqlite3
import os
import json
import threading
import time
import socket
import calendar
from datetime import datetime, date
import customtkinter as ctk
from tkinter import messagebox, ttk
from dotenv import load_dotenv
from PIL import Image, ImageDraw
import numpy as np

from database import DB_PATH, ENV_PATH

# Carrega o .env da pasta de dados segura do usuário
_ENV_PATH = ENV_PATH
load_dotenv(dotenv_path=_ENV_PATH, override=True)
from sync_manager import inicializar_fila_local, iniciar_thread_sync, enfileirar_log, enfileirar_sync
from falta_automatica import iniciar_robo_faltas
from calendar_engine import get_disciplina_atual
import auth

# ──────────────────────────────────────────────────────────
# ÍCONE DO SAPO — carregado uma vez, reutilizado em toda a UI
# ──────────────────────────────────────────────────────────
def _carregar_icone_sapo(tamanho=(48, 48)):
    """Carrega sapa_icon_white.png ao lado do main.py. Fallback para emoji se não achar."""
    try:
        _path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sapa_icon_white.png")
        pil_img = Image.open(_path).convert("RGBA")
        return ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=tamanho)
    except Exception:
        return None

# Cache dos ícones em tamanhos padrão
_ICONES: dict = {}

# ──────────────────────────────────────────────────────────
# TEMA E CORES
# ──────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

COR_VERDE       = "#1DB954"
COR_VERMELHO    = "#E53935"
COR_LARANJA     = "#FF8C00"
COR_AZUL        = "#34a853"  # Tom de verde do registro de presença
COR_ROXO        = "#7B2FBE"
COR_MUTED       = "#8888aa"
COR_CARD        = "#1e1e2e"
COR_CARD2       = "#2a2a3d"

DIAS_PT = ["Segunda-feira", "Terça-feira", "Quarta-feira",
           "Quinta-feira",  "Sexta-feira", "Sábado",       "Domingo"]
MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
DIAS_SEMANA  = DIAS_PT
MAPA_DIAS    = {d: i for i, d in enumerate(DIAS_PT)}
INV_MAPA_DIAS = {i: d for d, i in MAPA_DIAS.items()}

MATRIZ_CURRICULAR = [
    "ALGORITMO E LOGICA DE PROGRAMACAO","BANCO DE DADOS","ENGENHARIA DE REQUISITOS",
    "EXTENSAO ACADEMICA","PROGRAMACAO ESTRUTURADA","DESENVOLVIMENTO ORIENTADO A OBJETOS",
    "DESENVOLVIMENTO PARA AMBIENTES MOVEIS","DESENVOLVIMENTO WEB","ENGENHARIA DE SOFTWARE",
    "FUNDAMENTOS E ARQUITETURA DE COMPUTADORES","REDES DE COMPUTADORES","SISTEMAS OPERACIONAIS",
    "ADMINISTRACAO DE REDES DE COMPUTADORES","PLANEJAMENTO E PROJETO DE REDES DE COMPUTADORES",
    "TOPICOS AVANCADOS EM SEGURANCA DA INFORMACAO","EMPREENDEDORISMO",
    "ETICA E RESPONSABILIDADE SOCIO AMBIENTAL","MODELAGEM E GESTAO DE PROCESSOS",
    "GERENCIAMENTO DE PROJETOS","GERENCIAMENTO DE SERVICOS DE TI",
    "GOVERNANCA CORPORATIVA E DE TI","LIBRAS",
]

# ──────────────────────────────────────────────────────────
# ESTADO GLOBAL
# ──────────────────────────────────────────────────────────
DISCIPLINA_SELECIONADA = None
ID_DISCIPLINA_ATUAL    = None
ADMIN_LOGADO           = False   # login persistente durante a sessão

# ── Buffer do leitor USB (qualquer leitor que emule teclado) ──
_kb_buffer    = []      # chars acumulados entre teclas
_kb_ultimo    = ""      # último RA processado (debounce)
_kb_ultimo_ts = 0.0    # timestamp do último processamento

# ── Wi-Fi ──
_online = False

# ── Referências de widgets para funções chamadas antes da UI estar pronta ──
_ui_refs: dict = {}

# ──────────────────────────────────────────────────────────
# BANCO
# ──────────────────────────────────────────────────────────
def conectar():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def garantir_schema():
    mig = [
        ("PROFESSORES","Email","TEXT"),
        ("PROFESSORES","Telefone","TEXT"),
        ("PROFESSORES","senha_hash","TEXT"),
    ]
    try:
        with conectar() as conn:
            for t, c, tp in mig:
                try:
                    conn.execute(f"ALTER TABLE {t} ADD COLUMN {c} {tp}")
                except sqlite3.OperationalError:
                    pass
            conn.commit()
    except Exception:
        pass

# ──────────────────────────────────────────────────────────
# HELPERS — CENTRALIZAR JANELA
# ──────────────────────────────────────────────────────────
def centralizar(janela, largura, altura, pai=None):
    """Centraliza uma CTkToplevel na tela (ou na janela pai) e a traz para a frente."""
    janela.update_idletasks()
    if pai:
        px = pai.winfo_x() + pai.winfo_width()  // 2
        py = pai.winfo_y() + pai.winfo_height() // 2
        janela.transient(pai)
    else:
        px = janela.winfo_screenwidth()  // 2
        py = janela.winfo_screenheight() // 2
        janela.transient(root)

    x = px - largura  // 2
    y = py - altura   // 2
    janela.geometry(f"{largura}x{altura}+{x}+{y}")

    # Força a janela a ficar sempre à frente da principal (Topmost)
    janela.after(100, lambda: janela.attributes("-topmost", True))
    janela.after(200, lambda: janela.focus_force())

# ──────────────────────────────────────────────────────────
# HELPERS — STATUS
# ──────────────────────────────────────────────────────────
def _mostrar_status(texto, cor):
    status_label.configure(text=texto, text_color=cor)
    root.after(6000, lambda: status_label.configure(
        text="Aponte a carteirinha no leitor", text_color=COR_MUTED))

# ──────────────────────────────────────────────────────────
# LIMPEZA DO RA
# ──────────────────────────────────────────────────────────
def limpar_ra(ra_bruto):
    """
    Extrai o RA do padrão da carteirinha UNISEPE.
    Casos tratados:
      "2026 15652"  → "15652"  (com espaço: pega após o espaço)
      "202615652"   → "15652"  (colado 9+ dígitos: descarta os 4 primeiros = ano)
      "15652"       → "15652"  (já limpo: retorna como está)
    """
    texto = str(ra_bruto).strip()
    if not texto:
        return ""
    if " " in texto:
        return texto.split()[-1].strip()
    digitos = "".join(c for c in texto if c.isdigit())
    if len(digitos) >= 9:
        return digitos[4:]
    return digitos if digitos else texto

def _buscar_id_disciplina(nome, cursor):
    cursor.execute(
        "SELECT ID FROM DISCIPLINAS "
        "WHERE Nome_Materia||' - '||Semestre||' ('||Bloco||') - '||Professor_Nome = ?",
        (nome,))
    r = cursor.fetchone()
    return r["ID"] if r else None

# ──────────────────────────────────────────────────────────
# REGISTRO DE PRESENÇA  (registro único por aluno por dia)
# ──────────────────────────────────────────────────────────
def registrar_presenca(ra_bruto):
    global DISCIPLINA_SELECIONADA, ID_DISCIPLINA_ATUAL

    if not DISCIPLINA_SELECIONADA:
        _mostrar_status("⚠️  Nenhuma aula selecionada!", COR_LARANJA)
        return

    ra = limpar_ra(ra_bruto)
    if not ra:
        return

    agora      = datetime.now()
    data_hj    = agora.strftime("%d/%m/%Y")
    hora_atual = agora.strftime("%H:%M:%S")

    with conectar() as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT Nome FROM ALUNOS WHERE RA=?", (ra,))
        aluno = cursor.fetchone()
        if not aluno:
            # autocadastro precisa rodar na thread da UI
            root.after(0, lambda: _abrir_autocadastro(ra))
            return

        nome = aluno["Nome"]

        # Busca registro único do dia (uma linha por aluno por disciplina por dia)
        cursor.execute(
            "SELECT ID, Hora_Entrada, Hora_Saida FROM LOGS "
            "WHERE RA_Aluno=? AND Data=? AND Disciplina=?",
            (ra, data_hj, DISCIPLINA_SELECIONADA))
        reg = cursor.fetchone()

        id_disc = ID_DISCIPLINA_ATUAL or _buscar_id_disciplina(DISCIPLINA_SELECIONADA, cursor) or 0

        if reg is None:
            # Primeiro bip — cria linha com Entrada
            # Hora preenchida igual a Hora_Entrada para compatibilidade com schema (NOT NULL)
            cursor.execute(
                "INSERT INTO LOGS (RA_Aluno,Data,Hora,Hora_Entrada,Disciplina,Tipo) "
                "VALUES (?,?,?,?,?,'ENTRADA')",
                (ra, data_hj, hora_atual, hora_atual, DISCIPLINA_SELECIONADA))
            conn.commit()
            enfileirar_log(ra_aluno=int(ra), disciplina_id=id_disc, tipo="ENTRADA")
            root.after(0, lambda n=nome: _mostrar_status(f"✅  Presença Confirmada\n{n}", COR_VERDE))

        elif reg["Hora_Saida"] is None:
            # Já tem entrada, sem saída — trava 5 min
            hora_entrada_str = reg["Hora_Entrada"] or reg["Hora"] or hora_atual
            try:
                entrada_dt = datetime.strptime(f"{data_hj} {hora_entrada_str}", "%d/%m/%Y %H:%M:%S")
            except ValueError:
                entrada_dt = agora  # fallback seguro: libera registro de saída
            delta = (agora - entrada_dt).total_seconds()
            if delta < 300:
                restam = int((300 - delta) / 60) + 1
                root.after(0, lambda n=nome, r=restam: _mostrar_status(
                    f"⏳  {n} — aguarde {r} min para registrar saída.", COR_LARANJA))
                return
            # Atualiza a mesma linha com a hora de saída (Hora também para o robô de faltas)
            cursor.execute(
                "UPDATE LOGS SET Hora_Saida=?, Hora=?, Tipo='SAIDA' WHERE ID=?",
                (hora_atual, hora_atual, reg["ID"]))
            conn.commit()
            enfileirar_log(ra_aluno=int(ra), disciplina_id=id_disc, tipo="SAIDA")
            root.after(0, lambda n=nome: _mostrar_status(f"👋  Saída Confirmada\n{n}", COR_AZUL))

        else:
            root.after(0, lambda n=nome: _mostrar_status(
                f"✅  {n} já tem entrada e saída registradas.", COR_AZUL))

# ──────────────────────────────────────────────────────────
# LEITOR USB — buffer de teclado global
# Qualquer leitor que emule teclado funciona: digita os chars
# e manda <Return>. root.bind("<Key>") captura tudo.
# O Entry manual usa "break" para não duplicar o evento.
# ──────────────────────────────────────────────────────────
def _on_key_global(event):
    """Intercepta teclas do root. Chars → buffer; Return → processa RA."""
    global _kb_buffer, _kb_ultimo, _kb_ultimo_ts
    ch  = event.char
    sym = event.keysym

    if sym == "Return":
        ra_bruto = "".join(_kb_buffer).strip()
        _kb_buffer.clear()
        if not ra_bruto:
            return
        # Rejeita silenciosamente códigos que não são RA da UNISEPE
        if not _ra_parece_valido(ra_bruto):
            return
        ra_limpo = limpar_ra(ra_bruto)
        agora_ts = time.time()
        if ra_limpo == _kb_ultimo and (agora_ts - _kb_ultimo_ts) < 2:
            return
        _kb_ultimo    = ra_limpo
        _kb_ultimo_ts = agora_ts
        registrar_presenca(ra_limpo)
    elif sym in ("BackSpace", "Delete"):
        # Leitor USB nunca manda Backspace, mas teclado manual pode
        if _kb_buffer:
            _kb_buffer.pop()
    elif sym == "Escape":
        # Cancela leitura em andamento
        _kb_buffer.clear()
    elif ch and ch.isprintable():
        _kb_buffer.append(ch)

def _on_entry_ra_return(event):
    """Fallback manual: usuário digita RA no campo e aperta Enter."""
    texto = ra_entry.get().strip()
    if texto:
        registrar_presenca(limpar_ra(texto))
        ra_entry.delete(0, "end")
    return "break"   # impede subir ao root e duplicar

# ──────────────────────────────────────────────────────────
# VALIDAÇÃO DE RA — só aceita padrão UNISEPE
# Padrão: 4 dígitos de ano (20xx) + 4-6 dígitos de RA
# Ex: "202615652", "2026 15652", "15652" (já limpo)
# ──────────────────────────────────────────────────────────
def _ra_parece_valido(ra_bruto: str) -> bool:
    """
    Filtra leituras inválidas antes de processar.
    Rejeita: strings muito curtas, muito longas, com letras, ou que não
    batem com o padrão ANO+RA da UNISEPE.
    """
    texto = str(ra_bruto).strip().replace(" ", "")
    # Só dígitos
    if not texto.isdigit():
        return False
    # RA limpo tem 4-6 dígitos; com ano tem 8-10 dígitos
    tamanho = len(texto)
    if tamanho < 4 or tamanho > 10:
        return False
    # Se tiver 8+ dígitos, os 4 primeiros devem ser um ano razoável (20xx)
    if tamanho >= 8:
        ano = int(texto[:4])
        if ano < 2000 or ano > 2099:
            return False
    return True

# ──────────────────────────────────────────────────────────
# PAINEL DO LEITOR — detecção de dispositivos USB + teste
# ──────────────────────────────────────────────────────────
def _listar_dispositivos_hid():
    """
    Lista dispositivos HID conectados, priorizando leitores de código de barras.
    Retorna lista de strings descritivas.
    """
    dispositivos = []
    leitores     = []   # separados para mostrar em destaque

    KEYWORDS_LEITOR = ["barcode","scanner","bar code","honeywell","zebra",
                       "datalogic","symbol","metrologic","opticon","newland",
                       "unitech","socket","denso","cognex","keyence","code reader",
                       "hid pos","pos hid","usb hid scanner"]
    KEYWORDS_IGNORAR = ["mouse","teclado","keyboard","touchpad","trackpad",
                        "gamepad","joystick","headset","audio","webcam","camera",
                        "biometric","fingerprint","hub","composite"]

    try:
        import subprocess, sys
        if sys.platform == "win32":
            resultado = subprocess.run(
                ["powershell", "-Command",
                 "Get-PnpDevice -Class HIDClass | Where-Object {$_.Status -eq 'OK'} | "
                 "Select-Object -ExpandProperty FriendlyName"],
                capture_output=True, text=True, timeout=5
            )
            for linha in resultado.stdout.splitlines():
                nome = linha.strip()
                if not nome:
                    continue
                nome_lower = nome.lower()
                # Pula dispositivos conhecidamente não-leitores
                if any(k in nome_lower for k in KEYWORDS_IGNORAR):
                    continue
                # Classifica como leitor se bater com keywords específicas
                if any(k in nome_lower for k in KEYWORDS_LEITOR):
                    leitores.append(f"✅  {nome}  ← leitor detectado")
                else:
                    dispositivos.append(f"   {nome}")
        else:
            # Linux
            try:
                with open("/proc/bus/input/devices") as f:
                    bloco = ""
                    for linha in f:
                        bloco += linha
                        if linha.strip() == "" and "Name" in bloco:
                            nl = [l for l in bloco.splitlines() if l.startswith("N: Name=")]
                            if nl:
                                nome = nl[0].replace('N: Name=','').strip().strip('"')
                                n_low = nome.lower()
                                if any(k in n_low for k in KEYWORDS_IGNORAR):
                                    pass
                                elif any(k in n_low for k in KEYWORDS_LEITOR):
                                    leitores.append(f"✅  {nome}  ← leitor detectado")
                                else:
                                    dispositivos.append(f"   {nome}")
                            bloco = ""
            except Exception:
                pass
    except Exception:
        pass

    resultado_final = leitores + dispositivos
    if not resultado_final:
        resultado_final = ["Nenhum dispositivo HID detectado. Verifique o cabo USB."]
    return resultado_final, bool(leitores)

def abrir_painel_leitor():
    jan = ctk.CTkToplevel(root)
    jan.title("🔌  Leitor de Código de Barras")
    jan.protocol("WM_DELETE_WINDOW", jan.destroy)
    jan.grab_set()
    centralizar(jan, 560, 540, root)

    # ── Cabeçalho ──
    ctk.CTkLabel(jan, text="🔌  Leitor de Código de Barras",
                 font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(18,4))
    ctk.CTkLabel(jan,
                 text="Qualquer leitor USB que emule teclado funciona automaticamente.\n"
                      "Conecte o cabo e bipe — sem instalação de driver necessária.",
                 font=ctk.CTkFont(size=11), text_color=COR_MUTED,
                 justify="center").pack(pady=(0,10))

    # ── Card: status do leitor ──
    card_status = ctk.CTkFrame(jan, corner_radius=10)
    card_status.pack(fill="x", padx=24, pady=4)

    lbl_status_icon = ctk.CTkLabel(card_status,
                                   text="🟢  Captura ativa — leitor pronto para uso",
                                   font=ctk.CTkFont(size=13, weight="bold"),
                                   text_color=COR_VERDE)
    lbl_status_icon.pack(pady=12)

    # ── Card: dispositivos detectados ──
    card_disp = ctk.CTkFrame(jan, corner_radius=10)
    card_disp.pack(fill="x", padx=24, pady=4)

    frame_disp_topo = ctk.CTkFrame(card_disp, fg_color="transparent")
    frame_disp_topo.pack(fill="x", padx=12, pady=(10,4))
    ctk.CTkLabel(frame_disp_topo, text="Dispositivos USB detectados:",
                 font=ctk.CTkFont(size=12, weight="bold")).pack(side="left")

    btn_redetectar = ctk.CTkButton(frame_disp_topo, text="🔄 Atualizar",
                                   width=90, height=26,
                                   fg_color=COR_CARD2, hover_color=COR_AZUL,
                                   font=ctk.CTkFont(size=11))
    btn_redetectar.pack(side="right")

    lista_disp = ctk.CTkTextbox(card_disp, height=80, font=ctk.CTkFont(size=11, family="Consolas"))
    lista_disp.pack(fill="x", padx=12, pady=(0,10))

    def _popular_dispositivos():
        lista_disp.configure(state="normal")
        lista_disp.delete("1.0", "end")
        devs, tem_leitor = _listar_dispositivos_hid()
        for d in devs:
            lista_disp.insert("end", f"{d}\n")
        if tem_leitor:
            lbl_status_icon.configure(
                text="🟢  Leitor de código de barras detectado e pronto!",
                text_color=COR_VERDE)
        else:
            lbl_status_icon.configure(
                text="🟡  Nenhum leitor identificado — conecte o leitor e clique em Atualizar",
                text_color=COR_LARANJA)
        lista_disp.configure(state="disabled")

    btn_redetectar.configure(command=_popular_dispositivos)
    _popular_dispositivos()

    # ── Card: campo de teste ──
    card_teste = ctk.CTkFrame(jan, corner_radius=10)
    card_teste.pack(fill="x", padx=24, pady=8)

    ctk.CTkLabel(card_teste, text="Teste — bipe uma carteirinha:",
                 font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=14, pady=(10,4))

    ent_teste = ctk.CTkEntry(card_teste, width=400, height=46,
                             font=ctk.CTkFont(size=20), justify="center",
                             placeholder_text="Aguardando leitura...")
    ent_teste.pack(padx=14, pady=(0,8))
    ent_teste.focus()

    # Card de resultado
    card_res = ctk.CTkFrame(jan, corner_radius=8, height=54)
    card_res.pack(fill="x", padx=24, pady=0)
    card_res.pack_propagate(False)
    lbl_res = ctk.CTkLabel(card_res, text="",
                           font=ctk.CTkFont(size=12, weight="bold"), text_color=COR_MUTED)
    lbl_res.place(relx=0.5, rely=0.5, anchor="center")

    def testar(event=None):
        bruto = ent_teste.get().strip()
        if not bruto:
            return "break"

        # Valida padrão RA antes de qualquer coisa
        if not _ra_parece_valido(bruto):
            lbl_res.configure(
                text=f"❌  '{bruto}' não parece um RA válido da UNISEPE.",
                text_color=COR_VERMELHO)
            ent_teste.delete(0, "end")
            return "break"

        limpo = limpar_ra(bruto)
        with conectar() as conn:
            aluno = conn.execute(
                "SELECT Nome, Turma FROM ALUNOS WHERE RA=?", (limpo,)
            ).fetchone()

        if aluno:
            lbl_res.configure(
                text=f"✅  RA {limpo} — {aluno['Nome']} ({aluno['Turma']})",
                text_color=COR_VERDE)
            lbl_status_icon.configure(
                text="🟢  Leitor funcionando perfeitamente!",
                text_color=COR_VERDE)
        else:
            lbl_res.configure(
                text=f"⚠️  RA {limpo} lido corretamente — aluno não cadastrado.",
                text_color=COR_LARANJA)
            lbl_status_icon.configure(
                text="🟡  RA lido, mas não encontrado no banco.",
                text_color=COR_LARANJA)

        ent_teste.delete(0, "end")
        return "break"

    ent_teste.bind("<Return>", testar)
    ctk.CTkButton(card_teste, text="🔍  Testar leitura",
                  command=testar, fg_color=COR_AZUL, height=36).pack(pady=(0,10))

    # ── Rodapé ──
    ctk.CTkLabel(jan,
                 text="💡  O SAPA só aceita códigos no padrão UNISEPE (ano + RA numérico).\n"
                      "    Outros tipos de código de barras são ignorados automaticamente.",
                 font=ctk.CTkFont(size=10), text_color=COR_MUTED,
                 justify="center").pack(side="bottom", pady=10)

# ──────────────────────────────────────────────────────────
# MONITOR DE WI-FI / INTERNET
# ──────────────────────────────────────────────────────────
def _verificar_internet() -> bool:
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.settimeout(2)
    try:
        s.connect(("8.8.8.8", 53))
        return True
    except OSError:
        return False
    finally:
        s.close()  # sempre fecha, mesmo em erro

def _loop_wifi():
    global _online
    while True:
        status = _verificar_internet()
        if status != _online:
            _online = status
            root.after(0, _atualizar_indicador_wifi)
        time.sleep(10)

def _atualizar_indicador_wifi():
    if _online:
        lbl_wifi.configure(text="🟢 Online",  text_color=COR_VERDE)
    else:
        lbl_wifi.configure(text="🔴 Offline", text_color=COR_VERMELHO)

def iniciar_monitor_wifi():
    threading.Thread(target=_loop_wifi, daemon=True).start()
    # Primeira verificação imediata (após 0.5s para a UI já estar pronta)
    def _primeira_verificacao():
        global _online
        _online = _verificar_internet()
        root.after(0, _atualizar_indicador_wifi)
    root.after(500, lambda: threading.Thread(target=_primeira_verificacao, daemon=True).start())

# ──────────────────────────────────────────────────────────
# AUTOCADASTRO
# ──────────────────────────────────────────────────────────
def _abrir_autocadastro(ra_limpo):
    jan = ctk.CTkToplevel(root)
    jan.title("🐸  Novo Aluno — Cadastro Rápido")
    jan.grab_set()
    jan.protocol("WM_DELETE_WINDOW", jan.destroy)  # X fecha sem travar
    centralizar(jan, 440, 380, root)

    ctk.CTkLabel(jan, text="👤  Aluno não encontrado",
                 font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(18,4))

    # Card com o RA detectado em destaque
    card_ra = ctk.CTkFrame(jan, corner_radius=8, height=40)
    card_ra.pack(fill="x", padx=40, pady=(0,14))
    card_ra.pack_propagate(False)
    ctk.CTkLabel(card_ra, text=f"RA detectado: {ra_limpo}",
                 font=ctk.CTkFont(size=13, weight="bold"),
                 text_color=COR_LARANJA).place(relx=0.5, rely=0.5, anchor="center")

    ctk.CTkLabel(jan, text="Nome Completo:").pack(anchor="w", padx=40, pady=(0,2))
    ent_nome = ctk.CTkEntry(jan, width=360,
                            placeholder_text="Digite o nome completo do aluno")
    ent_nome.pack(padx=40)
    ent_nome.focus()

    ctk.CTkLabel(jan, text="Turma:").pack(anchor="w", padx=40, pady=(10,2))
    cb = ctk.CTkComboBox(jan, values=["A","B","C"],
                         width=360)
    cb.set("ADS 3 SEM - B")   # turma padrão mais comum
    cb.pack(padx=40)

    def salvar(event=None):
        nome = ent_nome.get().strip()
        if not nome:
            messagebox.showerror("Obrigatório", "Digite o nome do aluno.", parent=jan)
            return
        if not ra_limpo.isdigit():
            messagebox.showerror("RA inválido", "RA deve ser numérico.", parent=jan)
            return

        ra_int = int(ra_limpo)
        turma  = cb.get()

        with conectar() as conn:
            conn.execute("INSERT OR IGNORE INTO ALUNOS (RA,Nome,Turma) VALUES (?,?,?)",
                         (ra_int, nome, turma))
            conn.commit()

        enfileirar_sync("alunos", {"ra": ra_int, "nome": nome, "turma": turma})

        jan.destroy()

        # Registra a entrada automaticamente logo após o cadastro
        # (sem precisar bipar de novo)
        root.after(100, lambda: registrar_presenca(ra_limpo))

    def cancelar():
        jan.destroy()
        _mostrar_status("⚠️  Cadastro cancelado — presença não registrada.", COR_LARANJA)

    ent_nome.bind("<Return>", salvar)

    frame_btns = ctk.CTkFrame(jan, fg_color="transparent")
    frame_btns.pack(pady=20)
    ctk.CTkButton(frame_btns, text="CADASTRAR E CONFIRMAR PRESENÇA",
                  fg_color=COR_VERDE, hover_color="#17a348",
                  command=salvar, height=42, width=280).pack(pady=(0,8))
    ctk.CTkButton(frame_btns, text="Cancelar",
                  fg_color="transparent", hover_color=COR_CARD2,
                  text_color=COR_MUTED, command=cancelar, height=32, width=140).pack()

# ──────────────────────────────────────────────────────────
# CALENDÁRIO VISUAL PRÓPRIO (sem tkcalendar)
# ──────────────────────────────────────────────────────────
def abrir_calendario(callback_data, pai=None):
    """
    Abre um mini-calendário CTk. Ao clicar num dia, chama callback_data("DD/MM/AAAA").
    """
    hoje = date.today()
    estado = {"ano": hoje.year, "mes": hoje.month}

    win = ctk.CTkToplevel(pai or root, fg_color=("#e6e6e6", COR_CARD))
    win.title("Selecionar Data")
    win.grab_set()
    win.resizable(False, False)
    centralizar(win, 340, 320, pai or root)

    frame_nav = ctk.CTkFrame(win, fg_color="transparent")
    frame_nav.pack(fill="x", padx=10, pady=8)

    lbl_mes = ctk.CTkLabel(frame_nav, text="", font=ctk.CTkFont(size=14, weight="bold"), width=200, text_color=("black", "white"))
    lbl_mes.pack(side="left", expand=True)

    frame_grid = ctk.CTkFrame(win, fg_color="transparent")
    frame_grid.pack(padx=10, pady=4, fill="both", expand=True)

    def desenhar():
        for w in frame_grid.winfo_children():
            w.destroy()
        a, m = estado["ano"], estado["mes"]
        lbl_mes.configure(text=f"{MESES_PT[m-1]}  {a}")

        for i, d in enumerate(["Seg","Ter","Qua","Qui","Sex","Sáb","Dom"]):
            ctk.CTkLabel(frame_grid, text=d, width=40,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=COR_MUTED).grid(row=0, column=i, padx=2, pady=2)

        primeiro_dia, total_dias = calendar.monthrange(a, m)
        col = primeiro_dia  # 0=Seg
        lin = 1
        for dia in range(1, total_dias + 1):
            d_obj = date(a, m, dia)
            eh_hoje = (d_obj == hoje)
            btn = ctk.CTkButton(
                frame_grid, text=str(dia), width=38, height=32,
                fg_color=COR_AZUL if eh_hoje else "transparent",
                text_color=("white", "white") if eh_hoje else ("black", "white"),
                hover_color=COR_VERDE,
                font=ctk.CTkFont(size=12, weight="bold" if eh_hoje else "normal"),
                command=lambda dd=dia: [
                    callback_data(f"{dd:02d}/{estado['mes']:02d}/{estado['ano']}"),
                    win.destroy()
                ]
            )
            btn.grid(row=lin, column=col, padx=2, pady=2)
            col += 1
            if col == 7:
                col = 0; lin += 1

    def mes_anterior():
        if estado["mes"] == 1:
            estado["mes"] = 12; estado["ano"] -= 1
        else:
            estado["mes"] -= 1
        desenhar()

    def mes_seguinte():
        if estado["mes"] == 12:
            estado["mes"] = 1; estado["ano"] += 1
        else:
            estado["mes"] += 1
        desenhar()

    ctk.CTkButton(frame_nav, text="◀", width=34, command=mes_anterior,
                  fg_color="transparent", text_color=("black", "white")).pack(side="left")
    ctk.CTkButton(frame_nav, text="▶", width=34, command=mes_seguinte,
                  fg_color="transparent", text_color=("black", "white")).pack(side="right")
    desenhar()

# ──────────────────────────────────────────────────────────
# DISCIPLINA AUTOMÁTICA
# ──────────────────────────────────────────────────────────
def atualizar_disciplina_automatica():
    """
    Detecta a disciplina do dia com base em:
      - Dia_Semana da DISCIPLINAS == dia da semana atual
      - Data_Inicio <= hoje <= Data_Fim
    Atualiza DISCIPLINA_SELECIONADA e sincroniza o combobox.
    Reagenda a si mesma a cada 60 s.
    """
    global DISCIPLINA_SELECIONADA, ID_DISCIPLINA_ATUAL
    aula = get_disciplina_atual()
    hora = datetime.now().strftime("%H:%M")

    if aula and isinstance(aula, dict):
        nome_completo = aula.get("nome_completo") or (
            f"{aula.get('mat','')} - {aula.get('sem','')} "
            f"({aula.get('bl','')}) - {aula.get('prof','')}"
        )
        DISCIPLINA_SELECIONADA = nome_completo
        ID_DISCIPLINA_ATUAL    = aula.get("id")

        # Atualiza card de aula
        texto = (f"📚  {aula.get('mat','')}  |  {aula.get('bl','')}  "
                 f"|  {aula.get('sem','')}  |  Prof. {aula.get('prof','')}")
        label_aula_card.configure(text=texto, text_color=COR_VERDE)

        # Sincroniza o combobox para mostrar a disciplina detectada
        if "combo_aula_ref" in _ui_refs:
            cb = _ui_refs["combo_aula_ref"]
            # Garante que o valor está na lista
            valores_atuais = list(cb.cget("values")) if hasattr(cb, "cget") else []
            if nome_completo not in valores_atuais:
                valores_atuais.insert(0, nome_completo)
                cb.configure(values=valores_atuais)
            cb.set(nome_completo)
    else:
        DISCIPLINA_SELECIONADA = None
        ID_DISCIPLINA_ATUAL    = None
        label_aula_card.configure(
            text=f"Sem aula programada para hoje ({hora})",
            text_color=COR_MUTED)
        if "combo_aula_ref" in _ui_refs:
            _ui_refs["combo_aula_ref"].set("Selecione a disciplina")

    root.after(60_000, atualizar_disciplina_automatica)

def atualizar_menu_principal():
    """Recarrega os comboboxes de disciplina a partir do banco."""
    try:
        with conectar() as conn:
            aulas = [r[0] for r in conn.execute(
                "SELECT Nome_Materia||' - '||Semestre||' ('||Bloco||') - '||Professor_Nome "
                "FROM DISCIPLINAS ORDER BY Nome_Materia")]
        # Atualiza o combobox da UI principal se já existir
        if "combo_aula_ref" in _ui_refs:
            _ui_refs["combo_aula_ref"].configure(
                values=aulas if aulas else ["Nenhuma disciplina cadastrada"])
    except Exception:
        pass

def ao_selecionar_disciplina(escolha):
    global DISCIPLINA_SELECIONADA, ID_DISCIPLINA_ATUAL
    if not escolha or escolha in ("Selecione a disciplina", "Nenhuma disciplina cadastrada"):
        return
    DISCIPLINA_SELECIONADA = escolha
    ID_DISCIPLINA_ATUAL    = None
    if "label_aula_card" in _ui_refs:
        _ui_refs["label_aula_card"].configure(text=f"📚  {escolha}", text_color=COR_VERDE)

# ──────────────────────────────────────────────────────────
# LOGIN PERSISTENTE
# ──────────────────────────────────────────────────────────
# BACKDOOR_USER e BACKDOOR_SENHA são lidos DINAMICAMENTE dentro
# de pedir_login() para garantir que o .env já foi carregado.
# Não declarar no topo do arquivo evita o bug de variável vazia.

def pedir_login(destino):
    """
    Se o admin já logou nesta sessão, abre direto o painel.
    Caso contrário pede credenciais.
    """
    global ADMIN_LOGADO
    if ADMIN_LOGADO:
        if destino == "admin":         abrir_painel_admin()
        elif destino == "professor":   abrir_painel_professor()
        elif destino == "config":      abrir_configuracoes()
        return

    garantir_schema()
    jan = ctk.CTkToplevel(root)
    jan.title("🔒  Autenticação")
    jan.grab_set()
    jan.protocol("WM_DELETE_WINDOW", jan.destroy)
    centralizar(jan, 380, 400, root)

    ctk.CTkLabel(jan, text="🔒  Acesso Restrito",
                 font=ctk.CTkFont(size=20, weight="bold")).pack(pady=25)
    ctk.CTkLabel(jan, text="E-mail:").pack(anchor="w", padx=50)
    ent_email = ctk.CTkEntry(jan, width=280); ent_email.pack(pady=5); ent_email.focus()
    ctk.CTkLabel(jan, text="Senha:").pack(anchor="w", padx=50)
    ent_senha = ctk.CTkEntry(jan, width=280, show="*"); ent_senha.pack(pady=5)
    lbl_err = ctk.CTkLabel(jan, text="", text_color=COR_VERMELHO,
                           font=ctk.CTkFont(size=12)); lbl_err.pack(pady=4)

    def tentar(event=None):
        global ADMIN_LOGADO
        email = ent_email.get().strip()
        senha = ent_senha.get()

        # Lê as credenciais do .env agora (não do topo do arquivo)
        # Garante que qualquer mudança nas configurações seja refletida
        load_dotenv(dotenv_path=_ENV_PATH, override=True)
        backdoor_user  = os.getenv("ADMIN_USER", "admin")
        backdoor_senha = os.getenv("ADMIN_PASS", "")

        ok = (
            (email == backdoor_user and senha == backdoor_senha and backdoor_senha != "")
            or auth.verificar_login(email, senha)
        )
        if ok:
            ADMIN_LOGADO = True
            jan.destroy()
            if destino == "admin":         abrir_painel_admin()
            elif destino == "professor":   abrir_painel_professor()
            elif destino == "config":      abrir_configuracoes()
        else:
            lbl_err.configure(text="E-mail ou senha incorretos.")
            ent_senha.delete(0,"end")

    ent_senha.bind("<Return>", tentar)
    ctk.CTkButton(jan, text="Entrar", command=tentar,
                  fg_color=COR_VERDE, hover_color="#17a348", height=40).pack(pady=18)
    # Credencial de backdoor só visível para admins logados — removida da tela pública

# ──────────────────────────────────────────────────────────
# PAINEL DO PROFESSOR  (registro único, calendário, justificativa)
# ──────────────────────────────────────────────────────────
def abrir_painel_professor():
    from relatorio import enviar_relatorio_por_email

    # Migração: adiciona colunas Hora_Entrada / Hora_Saida se ainda não existem
    with conectar() as conn:
        for col in ["Hora_Entrada", "Hora_Saida"]:
            try:
                conn.execute(f"ALTER TABLE LOGS ADD COLUMN {col} TEXT")
            except sqlite3.OperationalError:
                pass
        conn.commit()

    jan = ctk.CTkToplevel(root)
    jan.title("👨‍🏫  Painel do Professor")
    jan.protocol("WM_DELETE_WINDOW", jan.destroy)
    jan.lift()  # Eleva a janela para o topo
    jan.focus_force()  # Força o foco na janela
    jan.grab_set()  # Captura eventos do mouse para foco
    centralizar(jan, 960, 640, root)

    # ── Barra superior ──
    frame_top = ctk.CTkFrame(jan, fg_color="transparent")
    frame_top.pack(fill="x", padx=14, pady=8)

    ctk.CTkLabel(frame_top, text="Disciplina:", width=80).pack(side="left")
    with conectar() as conn:
        disc_lista = [r[0] for r in conn.execute(
            "SELECT Nome_Materia||' - '||Semestre||' ('||Bloco||') - '||Professor_Nome FROM DISCIPLINAS")]
    cb_disc = ctk.CTkComboBox(frame_top, values=disc_lista or ["Nenhuma"], width=440)
    cb_disc.set(DISCIPLINA_SELECIONADA or (disc_lista[0] if disc_lista else "Nenhuma"))
    cb_disc.pack(side="left", padx=6)

    # Campo de data + botão calendário
    data_hoje = datetime.now().strftime("%d/%m/%Y")
    ent_data = ctk.CTkEntry(frame_top, width=100); ent_data.insert(0, data_hoje)
    ent_data.pack(side="left", padx=4)

    def abrir_cal():
        def receber(d):
            ent_data.delete(0,"end"); ent_data.insert(0, d)
            carregar()
        abrir_calendario(receber, jan)

    ctk.CTkButton(frame_top, text="📅", width=38, command=abrir_cal,
                  fg_color=COR_CARD2).pack(side="left", padx=2)

    # ── Treeview com registro único (Entrada + Saída na mesma linha) ──
    style = ttk.Style()
    style.theme_use("default")
    style.configure("P.Treeview", background="#2b2b2b", foreground="white",
                    fieldbackground="#2b2b2b", rowheight=34, font=("Inter",12))
    style.configure("P.Treeview.Heading", background="#333333", foreground="#e0e0e0",
                    font=("Inter",12,"bold"), borderwidth=0)
    style.map("P.Treeview", background=[("selected",COR_AZUL)])

    frame_tree = ctk.CTkFrame(jan)
    frame_tree.pack(fill="both", expand=True, padx=14, pady=4)

    cols = ("RA","Nome","Turma","Entrada","Saída","Status","Justificativa")
    tree = ttk.Treeview(frame_tree, columns=cols, show="headings", style="P.Treeview")
    tree.heading("RA",          text="RA");          tree.column("RA",          width=70,  anchor="center")
    tree.heading("Nome",        text="Aluno");        tree.column("Nome",        width=220)
    tree.heading("Turma",       text="Turma");        tree.column("Turma",       width=110, anchor="center")
    tree.heading("Entrada",     text="Entrada");      tree.column("Entrada",     width=75,  anchor="center")
    tree.heading("Saída",       text="Saída");        tree.column("Saída",       width=75,  anchor="center")
    tree.heading("Status",      text="Status");       tree.column("Status",      width=110, anchor="center")
    tree.heading("Justificativa",text="Justificativa");tree.column("Justificativa",width=190)

    sb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=sb.set)
    sb.pack(side="right", fill="y")
    tree.pack(fill="both", expand=True)

    def carregar():
        for i in tree.get_children(): tree.delete(i)
        disc = cb_disc.get()
        data = ent_data.get().strip() or data_hoje

        with conectar() as conn:
            # Todos os alunos
            alunos = {r["RA"]: r for r in conn.execute(
                "SELECT RA, Nome, Turma FROM ALUNOS ORDER BY Nome")}

            # Registros do dia
            logs = {}
            for r in conn.execute(
                "SELECT * FROM LOGS WHERE Data=? AND Disciplina=?", (data, disc)):
                logs[r["RA_Aluno"]] = r

        for ra, al in alunos.items():
            reg = logs.get(ra)
            if reg:
                entrada = reg["Hora_Entrada"] or reg["Hora"] or ""
                saida   = reg["Hora_Saida"]   or ""
                tipo    = reg["Tipo"]
                just    = reg["Justificativa"] if "Justificativa" in reg.keys() else ""
            else:
                entrada, saida, tipo, just = "", "", "FALTA", ""

            tag = {"ENTRADA":"entrada","SAIDA":"saida",
                   "JUSTIFICADO":"just","FALTA":"falta"}.get(tipo,"falta")
            tree.insert("","end",
                values=(ra, al["Nome"], al["Turma"], entrada, saida, tipo, just or ""),
                tags=(tag,))

        tree.tag_configure("entrada",   foreground="#1DB954")
        tree.tag_configure("saida",     foreground="#64B5F6")
        tree.tag_configure("just",      foreground="#FFB300")
        tree.tag_configure("falta",     foreground="#E53935")

    # Vincula o combobox de disciplina para recarregar automaticamente ao mudar o valor
    cb_disc.configure(command=lambda e: carregar())

    # ── Botões ──
    frame_btn = ctk.CTkFrame(jan, fg_color="transparent")
    frame_btn.pack(fill="x", padx=14, pady=6)

    def _ra_selecionado():
        sel = tree.selection()
        if not sel: return None, None
        vals = tree.item(sel[0])["values"]
        return vals[0], vals[5]   # RA, Status atual

    def justificar():
        ra, status = _ra_selecionado()
        if ra is None: return

        disc = cb_disc.get()
        data = ent_data.get().strip() or data_hoje

        # Abre modal para digitar o motivo
        jwin = ctk.CTkToplevel(jan)
        jwin.title("Justificar Falta / Atraso")
        jwin.grab_set()
        centralizar(jwin, 440, 320, jan)

        ctk.CTkLabel(jwin, text="📝  Justificativa",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(pady=14)

        # Tipo de justificativa
        ctk.CTkLabel(jwin, text="Tipo:").pack(anchor="w", padx=30)
        cb_tipo = ctk.CTkComboBox(jwin, width=380,
            values=["Falta justificada","Atraso justificado","Saída antecipada justificada"])
        cb_tipo.pack(padx=30, pady=4)

        ctk.CTkLabel(jwin, text="Motivo:").pack(anchor="w", padx=30)
        txt_motivo = ctk.CTkEntry(jwin, width=380, placeholder_text="Descreva o motivo...")
        txt_motivo.pack(padx=30, pady=4)
        txt_motivo.focus()

        def confirmar():
            motivo = txt_motivo.get().strip()
            tipo_j = cb_tipo.get()
            if not motivo:
                messagebox.showerror("Obrigatório","Digite o motivo.", parent=jwin); return
            if not messagebox.askyesno(
                "Confirmar justificativa",
                f"Justificar como '{tipo_j}' para RA {ra}?\n\nMotivo: {motivo}",
                parent=jwin):
                return

            # Garante que a coluna Justificativa existe
            with conectar() as conn:
                try: conn.execute("ALTER TABLE LOGS ADD COLUMN Justificativa TEXT")
                except sqlite3.OperationalError: pass
                conn.commit()

            with conectar() as conn:
                existe = conn.execute(
                    "SELECT ID FROM LOGS WHERE RA_Aluno=? AND Data=? AND Disciplina=?",
                    (ra, data, disc)).fetchone()
                if existe:
                    conn.execute(
                        "UPDATE LOGS SET Tipo='JUSTIFICADO', Justificativa=? WHERE ID=?",
                        (f"[{tipo_j}] {motivo}", existe["ID"]))
                else:
                    # Cria registro de falta justificada mesmo sem bip
                    hora_agora = datetime.now().strftime("%H:%M:%S")
                    conn.execute(
                        "INSERT INTO LOGS (RA_Aluno,Data,Hora,Hora_Entrada,Disciplina,Tipo,Justificativa) "
                        "VALUES (?,?,?,?,?,'JUSTIFICADO',?)",
                        (ra, data, hora_agora, hora_agora, disc, f"[{tipo_j}] {motivo}"))
                conn.commit()
            jwin.destroy()
            carregar()

        ctk.CTkButton(jwin, text="✅  Confirmar Justificativa",
                      fg_color=COR_VERDE, command=confirmar, height=38).pack(pady=16)

    def remover_log():
        ra, _ = _ra_selecionado()
        if ra is None: return
        disc = cb_disc.get(); data = ent_data.get().strip() or data_hoje
        if not messagebox.askyesno("Confirmar","Remover registro deste aluno?", parent=jan): return
        with conectar() as conn:
            conn.execute("DELETE FROM LOGS WHERE RA_Aluno=? AND Data=? AND Disciplina=?",
                         (ra, data, disc)); conn.commit()
        carregar()

    def abrir_modal_selecao_professor(callback_enviar):
        pwin = ctk.CTkToplevel(jan)
        pwin.title("Selecione o Professor")
        pwin.grab_set()
        centralizar(pwin, 400, 250, jan)

        ctk.CTkLabel(pwin, text="👨‍🏫  Selecione o Professor", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=20)

        with conectar() as conn:
            profs = conn.execute("SELECT Nome_Professor, Email FROM PROFESSORES ORDER BY Nome_Professor").fetchall()

        if not profs:
            ctk.CTkLabel(pwin, text="Nenhum professor cadastrado.", text_color=COR_VERMELHO).pack()
            return

        lista_profs = [f"{p['Nome_Professor']} ({p['Email'] or 'Sem e-mail'})" for p in profs]
        cb_prof = ctk.CTkComboBox(pwin, values=lista_profs, width=320)
        cb_prof.pack(pady=10)

        def confirmar():
            idx = lista_profs.index(cb_prof.get())
            prof_selecionado = profs[idx]
            email = prof_selecionado["Email"]
            nome = prof_selecionado["Nome_Professor"]
            if not email:
                messagebox.showerror("Erro", f"O professor {nome} não possui e-mail cadastrado. Cadastre-o na Aba Admin.", parent=pwin)
                return
            pwin.destroy()
            callback_enviar(email, nome)

        ctk.CTkButton(pwin, text="Confirmar e Enviar", command=confirmar, fg_color=COR_VERDE).pack(pady=20)

    def baixar_excel_historico():
        """Excel completo de TODOS os registros (matriz) dividido em Abas por Bloco para Power BI."""
        from tkinter import filedialog
        from relatorio import extrair_infos_disciplina, inferir_periodo
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            messagebox.showerror("Erro", "O módulo 'openpyxl' não está instalado.\nAbra o terminal e digite: pip install openpyxl", parent=jan)
            return
        from datetime import datetime, timedelta

        destino = filedialog.asksaveasfilename(
            parent=jan,
            title="Salvar Excel — Histórico Completo por Blocos (Power BI)",
            defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            initialfile=f"sapa_HISTORICO_COMPLETO_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not destino:
            return  # usuário cancelou

        DIAS_PT = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]

        def _parse_data(texto):
            if not texto: return None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try: return datetime.strptime(texto.strip(), fmt).date()
                except ValueError: continue
            return None

        with conectar() as conn:
            conn.row_factory = sqlite3.Row
            alunos_rows = conn.execute("SELECT RA, Nome, Turma FROM ALUNOS").fetchall()
            disc_rows = conn.execute("SELECT * FROM DISCIPLINAS").fetchall()
            logs_rows = conn.execute("SELECT * FROM LOGS").fetchall()

        alunos = {r["RA"]: dict(r) for r in alunos_rows}

        logs_map = {}
        for l in logs_rows:
            chave = (l["Data"], l["RA_Aluno"], l["Disciplina"])
            if chave not in logs_map:
                logs_map[chave] = []
            logs_map[chave].append(dict(l))

        export_data = []
        logs_processados = set()

        # 1. Gerar Matriz a partir do Calendário
        for d in disc_rows:
            mat = d["Nome_Materia"] or ""
            sem = d["Semestre"] or ""
            blo = d["Bloco"] or ""
            prof = d["Professor_Nome"] or ""
            dia_sem = d["Dia_Semana"] or ""

            nome_completo = f"{mat} - {sem} ({blo}) - {prof}"

            d_ini = _parse_data(d["Data_Inicio"])
            d_fim = _parse_data(d["Data_Fim"])

            if d_ini and d_fim and dia_sem in DIAS_PT:
                dia_idx = DIAS_PT.index(dia_sem)
                curr_date = d_ini
                while curr_date <= d_fim:
                    if curr_date.weekday() == dia_idx:
                        data_str = curr_date.strftime("%d/%m/%Y")

                        for ra, aluno in alunos.items():
                            chave = (data_str, ra, nome_completo)
                            if chave in logs_map:
                                for l in logs_map[chave]:
                                    logs_processados.add(l["ID"])
                                    export_data.append({
                                        "sort_date": curr_date.strftime("%Y%m%d"),
                                        "Bloco": blo,
                                        "Materia": mat,
                                        "Nome_Aluno": aluno["Nome"],
                                        "ID": l["ID"],
                                        "RA": ra,
                                        "Turma": aluno["Turma"],
                                        "Data": data_str,
                                        "Hora_Entrada": l["Hora_Entrada"] or "",
                                        "Hora_Saida": l["Hora_Saida"] or "",
                                        "Professor": prof,
                                        "Semestre": sem,
                                        "Status": l["Tipo"],
                                        "Justificativa": l["Justificativa"] or ""
                                    })
                            else:
                                export_data.append({
                                    "sort_date": curr_date.strftime("%Y%m%d"),
                                    "Bloco": blo,
                                    "Materia": mat,
                                    "Nome_Aluno": aluno["Nome"],
                                    "ID": "-",
                                    "RA": ra,
                                    "Turma": aluno["Turma"],
                                    "Data": data_str,
                                    "Hora_Entrada": "",
                                    "Hora_Saida": "",
                                    "Professor": prof,
                                    "Semestre": sem,
                                    "Status": "FALTA",
                                    "Justificativa": ""
                                })
                    curr_date += timedelta(days=1)

        # 2. Adicionar logs reais que ficaram de fora do calendário
        for l in logs_rows:
            if l["ID"] not in logs_processados:
                ra = l["RA_Aluno"]
                aluno = alunos.get(ra, {"Nome": "Desconhecido", "Turma": "-"})

                mat_e, sem_e, blo_e, prof_e = extrair_infos_disciplina(l["Disciplina"])
                d_sort = "99999999"
                pd = _parse_data(l["Data"])
                if pd: d_sort = pd.strftime("%Y%m%d")

                export_data.append({
                    "sort_date": d_sort,
                    "Bloco": blo_e,
                    "Materia": mat_e,
                    "Nome_Aluno": aluno["Nome"],
                    "ID": l["ID"],
                    "RA": ra,
                    "Turma": aluno["Turma"],
                    "Data": l["Data"],
                    "Hora_Entrada": l["Hora_Entrada"] or "",
                    "Hora_Saida": l["Hora_Saida"] or "",
                    "Professor": prof_e,
                    "Semestre": sem_e,
                    "Status": l["Tipo"],
                    "Justificativa": l["Justificativa"] or ""
                })

        # 3. Ordenação
        export_data.sort(key=lambda x: (x["Bloco"], x["sort_date"], x["Materia"], x["Nome_Aluno"]))

        if not export_data:
            messagebox.showwarning("Sem dados", "Nenhum registro ou cronograma válido encontrado.", parent=jan)
            return

        # 4. Agrupar e exportar para Excel (.xlsx) com abas
        wb = Workbook()
        wb.remove(wb.active)  # Remove aba padrão inicial

        blocos_dict = {}
        for d in export_data:
            b_name = d["Bloco"].strip()
            if not b_name:
                b_name = "Outros"
            if b_name not in blocos_dict:
                blocos_dict[b_name] = []
            blocos_dict[b_name].append(d)

        cabecalho = [
            "Log_ID", "RA", "Aluno_Nome", "Turma", "Periodo_Inferido",
            "Data", "Hora_Entrada", "Hora_Saida",
            "Materia", "Professor", "Semestre", "Bloco",
            "Status", "Justificativa"
        ]

        for b_name, linhas in blocos_dict.items():
            safe_title = b_name[:31]  # Limite do Excel
            ws = wb.create_sheet(title=safe_title)

            ws.append(cabecalho)
            for col in range(1, len(cabecalho) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)

            for d in linhas:
                per = inferir_periodo(d["Turma"])
                ws.append([
                    d["ID"], d["RA"], d["Nome_Aluno"], d["Turma"], per,
                    d["Data"], d["Hora_Entrada"], d["Hora_Saida"],
                    d["Materia"], d["Professor"], d["Semestre"], d["Bloco"],
                    d["Status"], d["Justificativa"]
                ])

        try:
            wb.save(destino)
        except Exception as e:
            messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar o arquivo Excel.\nErro: {e}", parent=jan)
            return

        messagebox.showinfo(
            "✅  Excel Histórico Salvo",
            f"Arquivo Excel salvo em:\n{destino}\n\n"
            f"Total: {len(export_data)} registros organizados em {len(blocos_dict)} abas (Blocos).\n\n"
            "📌  Pronto para importar no Power BI ou Excel!",
            parent=jan)

    def baixar_csv_dia():
        """CSV apenas do dia selecionado no painel — inclui alunos com FALTA."""
        from tkinter import filedialog
        from relatorio import extrair_infos_disciplina, inferir_periodo
        import csv as _csv

        data_sel = ent_data.get().strip() or data_hoje
        disc_sel = cb_disc.get()

        destino = filedialog.asksaveasfilename(
            parent=jan,
            title=f"Salvar CSV — Dia {data_sel}",
            defaultextension=".csv",
            filetypes=[("CSV separado por ponto-e-vírgula", "*.csv"), ("Todos os arquivos", "*.*")],
            initialfile=f"sapa_{data_sel.replace('/','_')}_{disc_sel[:20].replace(' ','_')}.csv"
        )
        if not destino:
            return

        # Pega todos os alunos + logs do dia (mesma lógica da treeview)
        with conectar() as conn:
            conn.row_factory = sqlite3.Row
            alunos = list(conn.execute("SELECT RA, Nome, Turma FROM ALUNOS ORDER BY Nome"))
            logs_dia = {r["RA_Aluno"]: r for r in conn.execute(
                "SELECT * FROM LOGS WHERE Data=? AND Disciplina=?", (data_sel, disc_sel))}

        mat, sem, blo, prof = extrair_infos_disciplina(disc_sel)

        with open(destino, "w", newline="", encoding="utf-8-sig") as f:
            writer = _csv.writer(f, delimiter=";")
            writer.writerow([
                "RA", "Aluno_Nome", "Turma", "Periodo_Inferido",
                "Data", "Hora_Entrada", "Hora_Saida",
                "Materia", "Professor", "Semestre", "Bloco",
                "Status", "Justificativa"
            ])
            for al in alunos:
                reg = logs_dia.get(al["RA"])
                if reg:
                    entrada = reg["Hora_Entrada"] or reg["Hora"] or ""
                    saida   = reg["Hora_Saida"] or ""
                    status  = reg["Tipo"]
                    just    = reg["Justificativa"] or ""
                else:
                    entrada, saida, status, just = "", "", "FALTA", ""
                per = inferir_periodo(al["Turma"])
                writer.writerow([
                    al["RA"], al["Nome"], al["Turma"], per,
                    data_sel, entrada, saida,
                    mat, prof, sem, blo,
                    status, just
                ])

        messagebox.showinfo(
            "✅  CSV do Dia Salvo",
            f"Arquivo salvo em:\n{destino}\n\n"
            f"Data: {data_sel}  |  {len(alunos)} alunos registrados.",
            parent=jan)

    def enviar_email_dia():
        """Envia o relatório HTML do dia selecionado para o professor escolhido."""
        def executar(email_dest, nome_prof):
            from relatorio import enviar_relatorio_por_email
            data = ent_data.get().strip() or data_hoje
            enviar_relatorio_por_email(destinatario=email_dest, data_alvo=data, professor_alvo=nome_prof)
            messagebox.showinfo("E-mail", f"Relatório do dia {data}\nenviado para {email_dest}!", parent=jan)
        abrir_modal_selecao_professor(executar)

    def enviar_email_csv_completo():
        """Envia o CSV completo (todas as datas) por e-mail para Power BI."""
        def executar(email_dest, nome_prof):
            from relatorio import enviar_relatorio_completo_csv
            sucesso = enviar_relatorio_completo_csv(destinatario=email_dest, professor_alvo=nome_prof)
            if sucesso:
                messagebox.showinfo("E-mail", f"CSV completo enviado para {email_dest}!", parent=jan)
            else:
                messagebox.showerror("Erro", "Falha ao enviar. Verifique as credenciais no .env.", parent=jan)
        abrir_modal_selecao_professor(executar)

    # ── Linha de botões ──
    ctk.CTkButton(frame_btn, text="Atualizar",
                  command=carregar, width=110, fg_color="transparent", border_width=1, text_color=COR_AZUL).pack(side="left", padx=3)
    ctk.CTkButton(frame_btn, text="Justificar",
                  command=justificar, width=120, fg_color="transparent", border_width=1, text_color=COR_AZUL).pack(side="left", padx=3)
    ctk.CTkButton(frame_btn, text="Remover",
                  command=remover_log, width=110, fg_color="transparent", border_width=1, border_color=COR_VERMELHO, text_color=COR_VERMELHO, hover_color="#401515").pack(side="left", padx=3)

    # Separador visual
    ctk.CTkLabel(frame_btn, text="|", text_color=COR_MUTED).pack(side="left", padx=4)

    ctk.CTkButton(frame_btn, text="Histórico Completo\n(Power BI .xlsx)",
                  command=baixar_excel_historico,
                  fg_color="transparent", border_width=1, border_color="#059669", text_color="#059669", hover_color="#18362a",
                  width=175, height=44,
                  font=ctk.CTkFont(size=11)).pack(side="left", padx=3)
    ctk.CTkButton(frame_btn, text="CSV do Dia\n(Data Selecionada)",
                  command=baixar_csv_dia,
                  fg_color="transparent", border_width=1, border_color=COR_AZUL, text_color=COR_AZUL,
                  width=155, height=44,
                  font=ctk.CTkFont(size=11)).pack(side="left", padx=3)
    ctk.CTkButton(frame_btn, text="Email\nRelatório do Dia",
                  command=enviar_email_dia,
                  fg_color="transparent", border_width=1, border_color=COR_AZUL, text_color=COR_AZUL,
                  width=140, height=44,
                  font=ctk.CTkFont(size=11)).pack(side="right", padx=3)
    ctk.CTkButton(frame_btn, text="Email\nCSV PBI",
                  command=enviar_email_csv_completo,
                  fg_color="transparent", border_width=1, border_color=COR_AZUL, text_color=COR_AZUL,
                  width=140, height=44,
                  font=ctk.CTkFont(size=11)).pack(side="right", padx=3)

    carregar()

# ──────────────────────────────────────────────────────────
# PAINEL DE ADMINISTRAÇÃO
# ──────────────────────────────────────────────────────────
def abrir_painel_admin():
    jan = ctk.CTkToplevel(root)
    jan.title("⚙️  SAPA — Administração")
    jan.protocol("WM_DELETE_WINDOW", jan.destroy)
    jan.lift()  # Eleva a janela para o topo
    jan.focus_force()  # Força o foco na janela
    jan.grab_set()  # Captura eventos do mouse para foco
    centralizar(jan, 1000, 700, root)

    tabs = ctk.CTkTabview(jan, segmented_button_selected_color=COR_AZUL, segmented_button_selected_hover_color="#2b8c45")
    tabs.pack(fill="both", expand=True, padx=10, pady=10)
    aba_prof   = tabs.add("Professores")
    aba_disc   = tabs.add("Disciplinas")
    aba_alunos = tabs.add("Alunos")

    # ─── Estilos Treeview ───
    style = ttk.Style()
    style.theme_use("default")
    style.configure("A.Treeview", background="#2b2b2b", foreground="white",
                    fieldbackground="#2b2b2b", rowheight=34, font=("Inter",12))
    style.configure("A.Treeview.Heading", background="#333333", foreground="#e0e0e0",
                    font=("Inter",12,"bold"), borderwidth=0)
    style.map("A.Treeview", background=[("selected",COR_AZUL)])

    def _campo(par, lbl, show=None):
        ctk.CTkLabel(par, text=lbl).pack(anchor="w", padx=12)
        e = ctk.CTkEntry(par, width=300, show=show or ""); e.pack(padx=12, pady=4); return e

    # ══ ABA PROFESSORES ══
    ff = ctk.CTkFrame(aba_prof, width=340, fg_color=("#ffffff", "#2b2b2b"), corner_radius=10)
    ff.pack(side="left", fill="y", padx=10, pady=10)
    ff.pack_propagate(False)
    ctk.CTkLabel(ff, text="Dados do Professor", font=ctk.CTkFont(weight="bold")).pack(pady=10)
    en = _campo(ff,"Nome Completo:")
    ee = _campo(ff,"E-mail (login):")
    es = _campo(ff,"Senha:",show="*")
    et = _campo(ff,"Telefone / WhatsApp:")
    fl = ctk.CTkFrame(aba_prof, fg_color="transparent")
    fl.pack(side="right", fill="both", expand=True, padx=10, pady=10)
    tp = ttk.Treeview(fl, columns=("ID","Nome","Email","Tel"), show="headings", style="A.Treeview")
    for h,w in [("ID",35),("Nome",180),("Email",200),("Tel",130)]:
        tp.heading(h,text=h); tp.column(h,width=w)
    tp.pack(fill="both", expand=True, pady=5)
    ped = {"id": ""}

    def _load_profs():
        for i in tp.get_children(): tp.delete(i)
        with conectar() as conn:
            for r in conn.execute("SELECT ID,Nome_Professor,COALESCE(Email,''),COALESCE(Telefone,'') FROM PROFESSORES"):
                tp.insert("","end",values=tuple(r))

    def _salvar_prof():
        nome=en.get().strip(); email=ee.get().strip() or f"{nome.split()[0].lower()}@unisepe.com.br"
        tel=et.get().strip(); senha=es.get()
        if not nome: messagebox.showerror("Erro","Nome obrigatório.", parent=jan); return
        h = auth.gerar_hash(senha) if senha else auth.gerar_hash("12345")
        try:
            with conectar() as conn:
                if ped["id"]:
                    if senha: conn.execute("UPDATE PROFESSORES SET Nome_Professor=?,Email=?,Telefone=?,senha_hash=? WHERE ID=?",(nome,email,tel,h,ped["id"]))
                    else:     conn.execute("UPDATE PROFESSORES SET Nome_Professor=?,Email=?,Telefone=? WHERE ID=?",(nome,email,tel,ped["id"]))
                    messagebox.showinfo("OK","Professor atualizado!", parent=jan)
                else:
                    conn.execute("INSERT INTO PROFESSORES (Nome_Professor,Email,Telefone,senha_hash) VALUES (?,?,?,?)",(nome,email,tel,h))
                    messagebox.showinfo("OK","Professor cadastrado!", parent=jan)
                conn.commit()
            enfileirar_sync("professores",{"nome":nome,"email":email,"senha_hash":h})
            for e in [en,ee,et,es]: e.delete(0,"end")
            ped["id"]=""; btn_sp.configure(text="Salvar Professor")
            _load_profs(); atualizar_menu_principal()
        except sqlite3.IntegrityError:
            messagebox.showerror("Duplicado","E-mail já cadastrado.", parent=jan)

    def _editar_prof():
        sel=tp.selection()
        if not sel: return
        v=tp.item(sel[0])["values"]; ped["id"]=v[0]
        for e,val in zip([en,ee,et],[v[1],v[2],v[3]]):
            e.delete(0,"end"); e.insert(0, str(val) if str(val) != "None" else "")
        es.delete(0,"end"); btn_sp.configure(text="Atualizar")

    def _apagar_prof():
        sel=tp.selection()
        if not sel: return
        if messagebox.askyesno("Confirmar","Apagar professor?", parent=jan):
            with conectar() as conn:
                conn.execute("DELETE FROM PROFESSORES WHERE ID=?",(tp.item(sel[0])["values"][0],)); conn.commit()
            _load_profs(); atualizar_menu_principal()

    btn_sp = ctk.CTkButton(ff, text="Salvar Professor", fg_color=COR_VERDE, command=_salvar_prof)
    btn_sp.pack(pady=12)
    fb = ctk.CTkFrame(fl, fg_color="transparent"); fb.pack(fill="x", pady=4)
    ctk.CTkButton(fb,text="Editar",fg_color=COR_AZUL,width=120,command=_editar_prof).pack(side="left",padx=4)
    ctk.CTkButton(fb,text="Apagar",fg_color=COR_VERMELHO,width=100,command=_apagar_prof).pack(side="right",padx=4)
    _load_profs()

    # ══ ABA DISCIPLINAS ══
    fd = ctk.CTkScrollableFrame(aba_disc, width=340, fg_color=("#ffffff", "#2b2b2b"), corner_radius=10)
    fd.pack(side="left", fill="y", padx=10, pady=10)
    ctk.CTkLabel(fd, text="Nova Disciplina", font=ctk.CTkFont(weight="bold")).pack(pady=10)
    ctk.CTkLabel(fd, text="Matéria:").pack(anchor="w", padx=12)
    cb_mat = ctk.CTkComboBox(fd, values=MATRIZ_CURRICULAR, width=300); cb_mat.pack(padx=12, pady=4)
    with conectar() as conn:
        profs_db=[r[0] for r in conn.execute("SELECT Nome_Professor FROM PROFESSORES")]
    ctk.CTkLabel(fd, text="Professor:").pack(anchor="w", padx=12)
    cb_pd = ctk.CTkComboBox(fd, values=profs_db or ["Nenhum"], width=300); cb_pd.pack(padx=12, pady=4)
    ctk.CTkLabel(fd, text="Semestre:").pack(anchor="w", padx=12)
    cb_sem = ctk.CTkComboBox(fd, values=[f"{i}º Semestre" for i in range(1,7)], width=300); cb_sem.pack(padx=12, pady=4)
    ctk.CTkLabel(fd, text="Bloco:").pack(anchor="w", padx=12)
    cb_bl = ctk.CTkComboBox(fd, values=[f"Bloco {i}" for i in range(1,5)], width=300); cb_bl.pack(padx=12, pady=4)
    ctk.CTkLabel(fd, text="Data Início (DD/MM/AAAA):").pack(anchor="w", padx=12)
    ent_ini = ctk.CTkEntry(fd, width=300, placeholder_text="Ex: 01/02/2026"); ent_ini.pack(padx=12, pady=4)
    ctk.CTkLabel(fd, text="Data Fim (DD/MM/AAAA):").pack(anchor="w", padx=12)
    ent_fim = ctk.CTkEntry(fd, width=300, placeholder_text="Ex: 30/06/2026"); ent_fim.pack(padx=12, pady=4)
    ctk.CTkLabel(fd, text="Dia da Semana:").pack(anchor="w", padx=12)
    cb_dia_disc = ctk.CTkComboBox(fd, values=DIAS_SEMANA, width=300); cb_dia_disc.pack(padx=12, pady=4)

    fld = ctk.CTkFrame(aba_disc, fg_color="transparent")
    fld.pack(side="right", fill="both", expand=True, padx=10, pady=10)

    scrollx = ttk.Scrollbar(fld, orient="horizontal")
    scrollx.pack(side="bottom", fill="x")
    scrolly = ttk.Scrollbar(fld, orient="vertical")
    scrolly.pack(side="right", fill="y")

    td = ttk.Treeview(fld, columns=("ID","Mat","Prof","Sem","Bloco","Inicio","Fim","Dia"), show="headings", style="A.Treeview")
    td.configure(xscrollcommand=scrollx.set, yscrollcommand=scrolly.set)
    scrollx.configure(command=td.xview)
    scrolly.configure(command=td.yview)

    for h,w in [("ID",35),("Mat",200),("Prof",160),("Sem",80),("Bloco",60),("Inicio",80),("Fim",80),("Dia",110)]:
        td.heading(h,text=h); td.column(h,width=w, stretch=False)
    td.pack(fill="both", expand=True, pady=5)

    def _load_disc():
        for i in td.get_children(): td.delete(i)
        with conectar() as conn:
            for r in conn.execute("SELECT ID,Nome_Materia,Professor_Nome,Semestre,Bloco,COALESCE(Data_Inicio,''),COALESCE(Data_Fim,''),COALESCE(Dia_Semana,'') FROM DISCIPLINAS"):
                td.insert("","end",values=tuple(r))
        # Garante que as barras de rolagem se atualizem
        td.update_idletasks()

    disc_em_edicao = {"id": ""}

    def _salvar_disc():
        mat  = cb_mat.get(); prof = cb_pd.get()
        sem  = cb_sem.get(); bl   = cb_bl.get()
        ini  = ent_ini.get().strip()
        fim  = ent_fim.get().strip()
        dia  = cb_dia_disc.get()

        if not mat or not prof or not dia:
            messagebox.showerror("Obrigatório",
                "Matéria, Professor e Dia da Semana são obrigatórios.", parent=jan)
            return

        # Valida datas se preenchidas
        from calendar_engine import _parse_data
        if ini and not _parse_data(ini):
            messagebox.showerror("Data inválida",
                f"Data Início '{ini}' inválida.\nUse o formato DD/MM/AAAA.", parent=jan)
            return
        if fim and not _parse_data(fim):
            messagebox.showerror("Data inválida",
                f"Data Fim '{fim}' inválida.\nUse o formato DD/MM/AAAA.", parent=jan)
            return
        if ini and fim:
            d_ini = _parse_data(ini)
            d_fim = _parse_data(fim)
            if d_ini and d_fim and d_fim < d_ini:
                messagebox.showerror("Intervalo inválido",
                    "A Data Fim não pode ser anterior à Data Início.", parent=jan)
                return

        with conectar() as conn:
            if disc_em_edicao["id"]:
                conn.execute(
                    "UPDATE DISCIPLINAS SET Nome_Materia=?,Professor_Nome=?,Semestre=?,Bloco=?,Data_Inicio=?,Data_Fim=?,Dia_Semana=? WHERE ID=?",
                    (mat, prof, sem, bl, ini, fim, dia, disc_em_edicao["id"]))
                messagebox.showinfo("OK","Disciplina atualizada!", parent=jan)
                disc_em_edicao["id"] = ""
                btn_disc.configure(text="Salvar Disciplina")
            else:
                conn.execute(
                    "INSERT INTO DISCIPLINAS (Nome_Materia,Professor_Nome,Semestre,Bloco,Data_Inicio,Data_Fim,Dia_Semana) VALUES (?,?,?,?,?,?,?)",
                    (mat, prof, sem, bl, ini, fim, dia))
                messagebox.showinfo("OK","Disciplina salva!", parent=jan)
            conn.commit()
        _load_disc(); atualizar_menu_principal()
        # Força reavaliação do calendário imediatamente
        root.after(100, atualizar_disciplina_automatica)

    def _editar_disc():
        sel = td.selection()
        if not sel: return
        v = td.item(sel[0])["values"]
        disc_em_edicao["id"] = v[0]
        cb_mat.set(str(v[1])); cb_pd.set(str(v[2])); cb_sem.set(str(v[3])); cb_bl.set(str(v[4]))
        ent_ini.delete(0,"end"); ent_ini.insert(0, str(v[5]) if v[5] and str(v[5]) != "None" else "")
        ent_fim.delete(0,"end"); ent_fim.insert(0, str(v[6]) if v[6] and str(v[6]) != "None" else "")
        cb_dia_disc.set(v[7] if v[7] and str(v[7]) != "None" else DIAS_SEMANA[0])
        btn_disc.configure(text="Atualizar Disciplina")

    def _apagar_disc():
        sel = td.selection()
        if not sel: return
        if messagebox.askyesno("Confirmar","Apagar disciplina?", parent=jan):
            with conectar() as conn:
                conn.execute("DELETE FROM DISCIPLINAS WHERE ID=?",(td.item(sel[0])["values"][0],)); conn.commit()
            _load_disc(); atualizar_menu_principal()

    btn_disc = ctk.CTkButton(fd, text="Salvar Disciplina", fg_color=COR_VERDE, command=_salvar_disc)
    btn_disc.pack(pady=12)
    fbd = ctk.CTkFrame(fld, fg_color="transparent"); fbd.pack(fill="x", pady=4)
    ctk.CTkButton(fbd, text="Editar", fg_color=COR_AZUL, width=110, command=_editar_disc).pack(side="left", padx=4)
    ctk.CTkButton(fbd, text="Apagar", fg_color=COR_VERMELHO, command=_apagar_disc).pack(side="right", padx=4)
    _load_disc()
    # ══════════════════════════════════════════════
    # ABA ALUNOS — listagem, busca, edição, exclusão
    # Apenas administradores chegam aqui.
    # ══════════════════════════════════════════════
    frame_busca = ctk.CTkFrame(aba_alunos, fg_color="transparent")
    frame_busca.pack(fill="x", padx=10, pady=(10,4))

    ctk.CTkLabel(frame_busca, text="Buscar:", width=55).pack(side="left")
    ent_busca_aluno = ctk.CTkEntry(frame_busca, width=280,
                                   placeholder_text="Nome ou RA...")
    ent_busca_aluno.pack(side="left", padx=6)

    lbl_total_alunos = ctk.CTkLabel(frame_busca, text="",
                                    text_color=COR_MUTED,
                                    font=ctk.CTkFont(size=11))
    lbl_total_alunos.pack(side="right", padx=10)

    # Treeview de alunos
    frame_tree_al = ctk.CTkFrame(aba_alunos, fg_color="transparent")
    frame_tree_al.pack(fill="both", expand=True, padx=10, pady=4)

    ta = ttk.Treeview(frame_tree_al,
                      columns=("RA","Nome","Turma"),
                      show="headings", style="A.Treeview")
    ta.heading("RA",    text="RA");    ta.column("RA",    width=90,  anchor="center")
    ta.heading("Nome",  text="Nome");  ta.column("Nome",  width=340)
    ta.heading("Turma", text="Turma"); ta.column("Turma", width=160, anchor="center")
    sb_al = ttk.Scrollbar(frame_tree_al, orient="vertical", command=ta.yview)
    ta.configure(yscrollcommand=sb_al.set)
    sb_al.pack(side="right", fill="y")
    ta.pack(fill="both", expand=True)

    # Painel de edição inline
    frame_edit_al = ctk.CTkFrame(aba_alunos, corner_radius=8, fg_color=("#ffffff", "#2b2b2b"))
    frame_edit_al.pack(fill="x", padx=10, pady=4)

    ctk.CTkLabel(frame_edit_al, text="RA:").grid(row=0, column=0, padx=8, pady=8, sticky="e")
    ent_al_ra   = ctk.CTkEntry(frame_edit_al, width=100, state="disabled")
    ent_al_ra.grid(row=0, column=1, padx=4, pady=8, sticky="w")

    ctk.CTkLabel(frame_edit_al, text="Nome:").grid(row=0, column=2, padx=8, pady=8, sticky="e")
    ent_al_nome = ctk.CTkEntry(frame_edit_al, width=280)
    ent_al_nome.grid(row=0, column=3, padx=4, pady=8, sticky="w")

    ctk.CTkLabel(frame_edit_al, text="Turma:").grid(row=0, column=4, padx=8, pady=8, sticky="e")
    cb_al_turma = ctk.CTkComboBox(frame_edit_al,
                                  values=["A","B","C"],
                                  width=150)
    cb_al_turma.grid(row=0, column=5, padx=4, pady=8, sticky="w")

    aluno_em_edicao = {"ra": ""}

    def _load_alunos(filtro=""):
        for i in ta.get_children():
            ta.delete(i)
        with conectar() as conn:
            if filtro:
                rows = conn.execute(
                    "SELECT RA,Nome,Turma FROM ALUNOS "
                    "WHERE Nome LIKE ? OR CAST(RA AS TEXT) LIKE ? "
                    "ORDER BY Nome",
                    (f"%{filtro}%", f"%{filtro}%")
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT RA,Nome,Turma FROM ALUNOS ORDER BY Nome"
                ).fetchall()
        for r in rows:
            ta.insert("","end", values=(r[0], r[1], r[2]))
        total = len(rows)
        lbl_total_alunos.configure(
            text=f"{total} aluno{'s' if total != 1 else ''} listado{'s' if total != 1 else ''}")

    def _ao_buscar(event=None):
        _load_alunos(ent_busca_aluno.get().strip())

    ent_busca_aluno.bind("<KeyRelease>", _ao_buscar)

    def _carregar_edicao():
        sel = ta.selection()
        if not sel:
            return
        v = ta.item(sel[0])["values"]
        aluno_em_edicao["ra"] = v[0]
        ent_al_ra.configure(state="normal")
        ent_al_ra.delete(0,"end"); ent_al_ra.insert(0, str(v[0]))
        ent_al_ra.configure(state="disabled")
        ent_al_nome.delete(0,"end"); ent_al_nome.insert(0, str(v[1]))
        cb_al_turma.set(str(v[2]))
        btn_salvar_al.configure(text="Salvar Alteração")

    def _salvar_edicao_aluno():
        ra  = aluno_em_edicao["ra"]
        if not ra:
            messagebox.showwarning("Selecione", "Clique em um aluno na lista e depois em Editar.", parent=jan)
            return
        novo_nome  = ent_al_nome.get().strip()
        nova_turma = cb_al_turma.get()
        if not novo_nome:
            messagebox.showerror("Obrigatório", "Nome não pode ficar vazio.", parent=jan)
            return
        with conectar() as conn:
            conn.execute("UPDATE ALUNOS SET Nome=?, Turma=? WHERE RA=?",
                         (novo_nome, nova_turma, ra))
            conn.commit()
        aluno_em_edicao["ra"] = ""
        btn_salvar_al.configure(text="Salvar Aluno")
        ent_al_nome.delete(0,"end")
        ent_al_ra.configure(state="normal"); ent_al_ra.delete(0,"end"); ent_al_ra.configure(state="disabled")
        _load_alunos(ent_busca_aluno.get().strip())
        messagebox.showinfo("Atualizado", f"Dados de RA {ra} atualizados!", parent=jan)

    def _excluir_aluno():
        sel = ta.selection()
        if not sel:
            messagebox.showwarning("Selecione", "Selecione um aluno na lista antes de excluir.", parent=jan)
            return
        v       = ta.item(sel[0])["values"]
        ra      = v[0]
        nome    = v[1]
        turma   = v[2]

        # ── Confirmação 1: pergunta simples ──
        if not messagebox.askyesno(
            "Excluir aluno",
            f"Deseja excluir o aluno:\n\n"
            f"RA: {ra}\nNome: {nome}\nTurma: {turma}\n\n"
            "Esta ação também remove todos os registros de presença deste aluno.",
            parent=jan
        ):
            return

        # ── Confirmação 2: confirma digitando o RA ──
        win_confirm = ctk.CTkToplevel(jan)
        win_confirm.title("Confirmação final")
        win_confirm.grab_set()
        win_confirm.protocol("WM_DELETE_WINDOW", win_confirm.destroy)
        centralizar(win_confirm, 380, 220, jan)

        ctk.CTkLabel(win_confirm,
                     text=f"⚠️  Ação irreversível",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color=COR_VERMELHO).pack(pady=(18,4))
        ctk.CTkLabel(win_confirm,
                     text=f"Digite o RA  {ra}  para confirmar a exclusão:",
                     font=ctk.CTkFont(size=12)).pack(pady=4)
        ent_confirm = ctk.CTkEntry(win_confirm, width=160, justify="center",
                                   font=ctk.CTkFont(size=16))
        ent_confirm.pack(pady=8)
        ent_confirm.focus()

        def _confirmar_exclusao(event=None):
            digitado = ent_confirm.get().strip()
            if digitado != str(ra):
                ent_confirm.delete(0,"end")
                ent_confirm.configure(placeholder_text="RA incorreto!")
                return
            with conectar() as conn:
                conn.execute("DELETE FROM LOGS  WHERE RA_Aluno=?", (ra,))
                conn.execute("DELETE FROM ALUNOS WHERE RA=?",       (ra,))
                conn.commit()
            win_confirm.destroy()
            _load_alunos(ent_busca_aluno.get().strip())
            messagebox.showinfo("Excluído",
                f"Aluno {nome} (RA {ra}) excluído com sucesso.", parent=jan)

        ent_confirm.bind("<Return>", _confirmar_exclusao)
        ctk.CTkButton(win_confirm, text="CONFIRMAR EXCLUSÃO",
                      fg_color=COR_VERMELHO, hover_color="#b71c1c",
                      command=_confirmar_exclusao, height=38).pack(pady=4)

    # Botões da aba Alunos
    frame_btns_al = ctk.CTkFrame(aba_alunos, fg_color="transparent")
    frame_btns_al.pack(fill="x", padx=10, pady=(0,8))

    ctk.CTkButton(frame_btns_al, text="Editar Selecionado",
                  fg_color=COR_AZUL, command=_carregar_edicao, width=160).pack(side="left", padx=4)
    btn_salvar_al = ctk.CTkButton(frame_btns_al, text="Salvar Aluno",
                                   fg_color=COR_VERDE, command=_salvar_edicao_aluno, width=160)
    btn_salvar_al.pack(side="left", padx=4)
    ctk.CTkButton(frame_btns_al, text="Excluir Aluno",
                  fg_color=COR_VERMELHO, hover_color="#b71c1c",
                  command=_excluir_aluno, width=150).pack(side="right", padx=4)

    _load_alunos()


# ──────────────────────────────────────────────────────────
# CONFIGURAÇÕES (engrenagem — só admin)
# ──────────────────────────────────────────────────────────
def abrir_configuracoes():
    if not ADMIN_LOGADO:
        pedir_login("config"); return

    jan = ctk.CTkToplevel(root)
    jan.title("⚙️  Configurações do Sistema")
    jan.grab_set()
    jan.protocol("WM_DELETE_WINDOW", jan.destroy)
    centralizar(jan, 620, 560, root)

    ctk.CTkLabel(jan, text="⚙️  Configurações do Sistema",
                 font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(18,8))

    tabs_cfg = ctk.CTkTabview(jan, height=380)
    tabs_cfg.pack(fill="both", expand=True, padx=20, pady=0)

    aba_admin = tabs_cfg.add("🔑  Administrador")
    aba_email = tabs_cfg.add("E-mail")
    aba_cloud = tabs_cfg.add("☁️  Banco de Dados")

    # ══ ABA ADMINISTRADOR ══
    ctk.CTkLabel(aba_admin,
                 text="Credenciais de acesso de emergência ao sistema.",
                 font=ctk.CTkFont(size=11), text_color=COR_MUTED).pack(pady=(10,14))

    ctk.CTkLabel(aba_admin, text="Usuário (login de emergência):").pack(anchor="w", padx=20)
    e_user = ctk.CTkEntry(aba_admin, width=520, placeholder_text="ex: admin")
    e_user.insert(0, os.getenv("ADMIN_USER", "admin"))
    e_user.pack(padx=20, pady=(3,10))

    ctk.CTkLabel(aba_admin, text="Senha atual:").pack(anchor="w", padx=20)
    e_pass_atual = ctk.CTkEntry(aba_admin, width=520, show="*",
                                placeholder_text="Digite a senha atual para confirmar")
    e_pass_atual.pack(padx=20, pady=(3,10))

    ctk.CTkLabel(aba_admin, text="Nova senha:").pack(anchor="w", padx=20)
    e_pass_nova = ctk.CTkEntry(aba_admin, width=520, show="*",
                               placeholder_text="Digite a nova senha")
    e_pass_nova.pack(padx=20, pady=(3,10))

    ctk.CTkLabel(aba_admin, text="Confirmar nova senha:").pack(anchor="w", padx=20)
    e_pass_conf = ctk.CTkEntry(aba_admin, width=520, show="*",
                               placeholder_text="Repita a nova senha")
    e_pass_conf.pack(padx=20, pady=(3,6))

    lbl_admin_status = ctk.CTkLabel(aba_admin, text="", font=ctk.CTkFont(size=11))
    lbl_admin_status.pack()

    def salvar_admin():
        user     = e_user.get().strip()
        atual    = e_pass_atual.get()
        nova     = e_pass_nova.get()
        conf     = e_pass_conf.get()
        senha_certa = os.getenv("ADMIN_PASS", "")

        if not user:
            lbl_admin_status.configure(text="❌  Usuário não pode ficar vazio.", text_color=COR_VERMELHO); return
        # Valida senha atual
        if atual != senha_certa:
            lbl_admin_status.configure(text="❌  Senha atual incorreta.", text_color=COR_VERMELHO); return
        if nova and nova != conf:
            lbl_admin_status.configure(text="❌  As senhas novas não coincidem.", text_color=COR_VERMELHO); return
        if nova and len(nova) < 6:
            lbl_admin_status.configure(text="❌  A nova senha deve ter ao menos 6 caracteres.", text_color=COR_VERMELHO); return

        _salvar_env({"ADMIN_USER": user, "ADMIN_PASS": nova if nova else senha_certa})
        lbl_admin_status.configure(
            text="✅  Credenciais atualizadas com sucesso!" + (" Reinicie para aplicar a nova senha." if nova else ""),
            text_color=COR_VERDE)

    ctk.CTkButton(aba_admin, text="Salvar Credenciais",
                  fg_color=COR_VERDE, command=salvar_admin, height=38).pack(pady=10)

    # ══ ABA E-MAIL ══
    ctk.CTkLabel(aba_email,
                 text="Configure o Gmail usado para enviar relatórios aos professores.",
                 font=ctk.CTkFont(size=11), text_color=COR_MUTED).pack(pady=(10,14))

    ctk.CTkLabel(aba_email, text="E-mail remetente (Gmail):").pack(anchor="w", padx=20)
    e_rem = ctk.CTkEntry(aba_email, width=520, placeholder_text="sapa@gmail.com")
    e_rem.insert(0, os.getenv("EMAIL_REMETENTE", ""))
    e_rem.pack(padx=20, pady=(3,10))

    ctk.CTkLabel(aba_email, text="Senha de App do Gmail:").pack(anchor="w", padx=20)
    e_sen = ctk.CTkEntry(aba_email, width=520, show="*",
                         placeholder_text="Gerada em myaccount.google.com → Senhas de App")
    e_sen.insert(0, os.getenv("EMAIL_SENHA", ""))
    e_sen.pack(padx=20, pady=(3,10))

    ctk.CTkLabel(aba_email, text="E-mail destinatário padrão (coordenação):").pack(anchor="w", padx=20)
    e_dest = ctk.CTkEntry(aba_email, width=520, placeholder_text="coordenacao@unisepe.com.br")
    e_dest.insert(0, os.getenv("EMAIL_RELATORIO", ""))
    e_dest.pack(padx=20, pady=(3,6))

    lbl_email_status = ctk.CTkLabel(aba_email, text="", font=ctk.CTkFont(size=11))
    lbl_email_status.pack()

    def testar_email():
        import smtplib
        rem  = e_rem.get().strip()
        sen  = e_sen.get().strip()
        dest = e_dest.get().strip() or rem
        if not rem or not sen:
            lbl_email_status.configure(text="❌  Preencha e-mail e senha antes de testar.", text_color=COR_VERMELHO); return
        lbl_email_status.configure(text="⏳  Conectando...", text_color=COR_MUTED); jan.update()
        try:
            with smtplib.SMTP("smtp.gmail.com", 587) as s:
                s.starttls(); s.login(rem, sen)
            lbl_email_status.configure(text="✅  Conexão OK! Gmail autenticado com sucesso.", text_color=COR_VERDE)
        except Exception as ex:
            lbl_email_status.configure(text=f"❌  {ex}", text_color=COR_VERMELHO)

    def salvar_email():
        _salvar_env({"EMAIL_REMETENTE": e_rem.get().strip(),
                     "EMAIL_SENHA":     e_sen.get().strip(),
                     "EMAIL_RELATORIO": e_dest.get().strip()})
        lbl_email_status.configure(text="✅  Configurações de e-mail salvas!", text_color=COR_VERDE)

    frame_btn_email = ctk.CTkFrame(aba_email, fg_color="transparent")
    frame_btn_email.pack(pady=10)
    ctk.CTkButton(frame_btn_email, text="🔍  Testar Conexão",
                  fg_color=COR_AZUL, command=testar_email, width=180).pack(side="left", padx=6)
    ctk.CTkButton(frame_btn_email, text="Salvar",
                  fg_color=COR_VERDE, command=salvar_email, width=120).pack(side="left", padx=6)

    # ══ ABA BANCO DE DADOS ══
    _api_url = os.getenv("SUPABASE_URL", "")
    _proj_id = _api_url.replace("https://", "").split(".")[0] if _api_url else ""
    _dash_url = f"https://supabase.com/dashboard/project/{_proj_id}" if _proj_id else "https://supabase.com/dashboard"

    ctk.CTkLabel(aba_cloud,
                 text="Conexão com o Supabase (banco de dados em nuvem).",
                 font=ctk.CTkFont(size=11), text_color=COR_MUTED).pack(pady=(10,14))

    ctk.CTkLabel(aba_cloud, text="Chave API (service_role):").pack(anchor="w", padx=20)
    e_key = ctk.CTkEntry(aba_cloud, width=520, show="*")
    e_key.insert(0, os.getenv("SUPABASE_KEY", ""))
    e_key.pack(padx=20, pady=(3,10))

    # Link para o dashboard
    frame_link = ctk.CTkFrame(aba_cloud, corner_radius=8)
    frame_link.pack(fill="x", padx=20, pady=(0,10))
    ctk.CTkLabel(frame_link, text="Painel do Banco:",
                 font=ctk.CTkFont(size=11, weight="bold")).pack(anchor="w", padx=12, pady=(8,2))
    ctk.CTkLabel(frame_link, text=_dash_url,
                 text_color="#4da6ff", font=ctk.CTkFont(size=10, underline=True),
                 cursor="hand2").pack(anchor="w", padx=12, pady=(0,4))

    def _abrir_dash(e=None):
        import webbrowser; webbrowser.open(_dash_url)

    ctk.CTkButton(frame_link, text="🌐  Abrir no Navegador",
                  command=_abrir_dash, fg_color=COR_AZUL, height=28, width=200).pack(anchor="w", padx=12, pady=(0,10))

    lbl_cloud_status = ctk.CTkLabel(aba_cloud, text="", font=ctk.CTkFont(size=11))
    lbl_cloud_status.pack()

    def testar_cloud():
        import requests as _req
        url = _api_url; key = e_key.get().strip()
        if not url or not key:
            lbl_cloud_status.configure(text="❌  URL ou chave não configurada.", text_color=COR_VERMELHO); return
        lbl_cloud_status.configure(text="⏳  Testando...", text_color=COR_MUTED); jan.update()
        try:
            r = _req.get(f"{url}/rest/v1/", headers={"apikey": key, "Authorization": f"Bearer {key}"}, timeout=6)
            if r.status_code == 200:
                lbl_cloud_status.configure(text="✅  Supabase respondendo — conexão OK!", text_color=COR_VERDE)
            elif r.status_code == 401:
                lbl_cloud_status.configure(text="❌  Chave inválida (401). Verifique a API Key.", text_color=COR_VERMELHO)
            else:
                lbl_cloud_status.configure(text=f"⚠️  Status {r.status_code}", text_color=COR_LARANJA)
        except Exception as ex:
            lbl_cloud_status.configure(text=f"❌  {ex}", text_color=COR_VERMELHO)

    def salvar_cloud():
        _salvar_env({"SUPABASE_KEY": e_key.get().strip()})
        lbl_cloud_status.configure(text="✅  Chave salva! Reinicie para aplicar.", text_color=COR_VERDE)

    frame_btn_cloud = ctk.CTkFrame(aba_cloud, fg_color="transparent")
    frame_btn_cloud.pack(pady=8)
    ctk.CTkButton(frame_btn_cloud, text="🔍  Testar Conexão",
                  fg_color=COR_AZUL, command=testar_cloud, width=180).pack(side="left", padx=6)
    ctk.CTkButton(frame_btn_cloud, text="Salvar",
                  fg_color=COR_VERDE, command=salvar_cloud, width=120).pack(side="left", padx=6)

# ── Função auxiliar: salvar no .env ──────────────────────────────────────────
def _salvar_env(novos: dict):
    """Atualiza chaves no .env sem apagar as demais, e aplica em memória imediato."""
    env_path = _ENV_PATH
    linhas = []
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            linhas = f.readlines()
    atualizadas = set()
    novas_linhas = []
    for linha in linhas:
        if "=" in linha and not linha.strip().startswith("#"):
            chave = linha.split("=")[0].strip()
            if chave in novos:
                novas_linhas.append(f"{chave}={novos[chave]}\n")
                atualizadas.add(chave); continue
        novas_linhas.append(linha)
    for k, v in novos.items():
        if k not in atualizadas:
            novas_linhas.append(f"{k}={v}\n")
    with open(env_path, "w", encoding="utf-8") as f:
        f.writelines(novas_linhas)
    for k, v in novos.items():
        os.environ[k] = v

# ──────────────────────────────────────────────────────────
# FECHAR O PROGRAMA — envia relatório antes de sair
# ──────────────────────────────────────────────────────────
def ao_fechar_programa():
    """
    Intercepta o X da janela principal.
    Estratégia de descoberta do professor (em ordem):
      1. DISCIPLINA_SELECIONADA (já está na memória — mais confiável)
      2. get_disciplina_atual() via grade/horário (fallback se nada selecionado)
    """
    def _descobrir_prof_hoje():
        aula_dict = None

        # ── Estratégia 1: disciplina já selecionada na sessão ──
        if DISCIPLINA_SELECIONADA and DISCIPLINA_SELECIONADA != "Selecione a disciplina":
            # Monta dict a partir do nome formatado "Mat - Sem (Bloco) - Prof"
            with conectar() as conn:
                row = conn.execute(
                    "SELECT d.ID, d.Nome_Materia, d.Semestre, d.Bloco, d.Professor_Nome, "
                    "       p.Email "
                    "FROM DISCIPLINAS d "
                    "LEFT JOIN PROFESSORES p ON p.Nome_Professor = d.Professor_Nome "
                    "WHERE d.Nome_Materia||' - '||d.Semestre||' ('||d.Bloco||') - '||d.Professor_Nome = ?",
                    (DISCIPLINA_SELECIONADA,)
                ).fetchone()
            if row:
                return (
                    row["Professor_Nome"],
                    row["Email"],
                    DISCIPLINA_SELECIONADA
                )

        # ── Estratégia 2: calendar_engine (fallback por horário) ──
        aula_dict = get_disciplina_atual()
        if not aula_dict or not isinstance(aula_dict, dict):
            return None, None, None

        nome_prof = aula_dict.get("prof", "")
        if not nome_prof:
            return None, None, None

        with conectar() as conn:
            row = conn.execute(
                "SELECT Email FROM PROFESSORES WHERE Nome_Professor=?", (nome_prof,)
            ).fetchone()
        email_prof = row["Email"] if row and row["Email"] else None
        nome_disc  = (f"{aula_dict.get('mat','')} - {aula_dict.get('sem','')} "
                      f"({aula_dict.get('bl','')}) - {nome_prof}")
        return nome_prof, email_prof, nome_disc

    nome_prof, email_prof, nome_disc = _descobrir_prof_hoje()

    if nome_prof and email_prof:
        msg_dialogo = (
            f"Deseja enviar o relatório da aula de hoje\n"
            f"({nome_disc})\n\n"
            f"para o professor {nome_prof}\n"
            f"({email_prof})?"
        )
    elif nome_prof and not email_prof:
        # Professor encontrado mas sem e-mail cadastrado
        msg_dialogo = (
            f"Professor {nome_prof} encontrado, mas sem e-mail cadastrado.\n\n"
            "Cadastre o e-mail no painel de Administração para enviar relatórios.\n\n"
            "Deseja fechar o SAPA mesmo assim?"
        )
    else:
        msg_dialogo = "Deseja fechar o SAPA?"

    resposta = messagebox.askyesnocancel("Fechar SAPA", msg_dialogo)

    if resposta is None:   # Cancelar — mantém o programa aberto
        return

    if resposta is False:  # Não — fecha sem enviar relatório
        root.destroy()
        return

    # Sim — tenta enviar relatório e depois fecha
    if nome_prof and email_prof:
        from relatorio import enviar_relatorio_por_email
        data_hj = datetime.now().strftime("%d/%m/%Y")
        try:
            enviar_relatorio_por_email(
                destinatario=email_prof,
                data_alvo=data_hj,
                professor_alvo=nome_prof
            )
            messagebox.showinfo(
                "✅  Relatório Enviado",
                f"Relatório enviado para:\n{email_prof}\n\nO SAPA será fechado agora."
            )
        except Exception as e:
            messagebox.showerror(
                "Erro ao enviar",
                f"Não foi possível enviar o e-mail:\n{e}\n\nO SAPA será fechado mesmo assim."
            )

    root.destroy()
_tela_cheia = False
_tamanho_anterior = None

def alternar_tela_cheia(event=None):
    global _tela_cheia, _tamanho_anterior
    _tela_cheia = not _tela_cheia

    if _tela_cheia:
        # Salva o tamanho atual e vai para tela cheia
        _tamanho_anterior = root.geometry()
        root.attributes('-fullscreen', True)
    else:
        # Volta para o tamanho anterior ou normal
        root.attributes('-fullscreen', False)
        if _tamanho_anterior:
            root.geometry(_tamanho_anterior)
        else:
            root.state('normal')

def adaptar_resolucao():
    """Adapta a interface para diferentes resoluções de tela"""
    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()

    # Define um tamanho baseado na resolução da tela
    if largura_tela < 1024 or altura_tela < 768:
        # Telas pequenas
        nova_largura = min(760, largura_tela - 40)
        nova_altura = min(740, altura_tela - 100)
    elif largura_tela >= 1920 and altura_tela >= 1080:
        # Telas grandes (Full HD+)
        nova_largura = 1200
        nova_altura = 900
    else:
        # Telas médias
        nova_largura = 1000
        nova_altura = 800

    root.geometry(f"{nova_largura}x{nova_altura}")

# ──────────────────────────────────────────────────────────
# UI PRINCIPAL — GLASSMORPHISM & TEMAS
# ──────────────────────────────────────────────────────────
def construir_ui():
    global root, status_label, ra_entry, label_aula_card, lbl_wifi, _TEMA_ATUAL

    _TEMA_ATUAL = "dark"
    ctk.set_appearance_mode(_TEMA_ATUAL)

    root = ctk.CTk()
    root.title("🐸  SAPA v7.0 — UNISEPE ADS")
    root.resizable(True, True)
    root.protocol("WM_DELETE_WINDOW", ao_fechar_programa)
    try:
        _ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sapa.ico")
        if os.path.exists(_ico):
            root.iconbitmap(_ico)
    except Exception:
        pass

    root.update_idletasks()
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    w  = min(1000, sw - 80)
    h  = min(800, sh - 80)
    x  = (sw - w) // 2
    y  = (sh - h) // 2
    root.geometry(f"{w}x{h}+{x}+{y}")
    root.minsize(720, 640)

    root.bind("<F11>", alternar_tela_cheia)
    root.bind("<Escape>", lambda e: alternar_tela_cheia() if root.attributes("-fullscreen") else None)
    root.bind("<Key>", _on_key_global)

    # ── TEMAS BASEADOS NO CSS ──
    TEMAS = {
        "light": {
            "bg": "#c8dfc8",          # verde salvia medio — visível, agradavel
            "hdr": "#d8eeda", "hdr_bor": "#a8cca8",
            "card": "#e4f2e4", "card_bor": "#a8cca8",
            "card_hover": "#4cd964",
            "text": "#0d1f0d",        # quase preto — leitura clara
            "muted": "#2d5a2d",       # verde escuro — substituí cinza fraco
            "icon_frog": "#1a6b1a"    # verde escuro no header
        },
        "dark": {
            "bg": "#07171d",
            "hdr": "#0f1e14", "hdr_bor": "#1a1a1a",
            "card": "#142819", "card_bor": "#1a1a1a",
            "card_hover": "#4cd964",
            "text": "#f5f5f7", "muted": "#86868b",
            "icon_frog": "#6ee782"
        }
    }
    def _tc(key): return TEMAS[_TEMA_ATUAL][key]

    root.configure(fg_color=_tc("bg"))

    # ── APLICAR TEMA ──
    _theme_widgets = []
    def _registrar_card(widget, fg_key, bor_key):
        _theme_widgets.append((widget, fg_key, bor_key))

    def _aplicar_tema_completo():
        root.configure(fg_color=_tc("bg"))
        for widget, fg_key, bor_key in _theme_widgets:
            if fg_key:
                try: widget.configure(fg_color=_tc(fg_key))
                except: pass
            if bor_key:
                try: widget.configure(border_color=_tc(bor_key))
                except: pass

        lbl_titulo_hdr.configure(text_color=_tc("icon_frog"))
        lbl_sub_hdr.configure(text_color=_tc("muted"))
        lbl_wifi.configure(text_color=_tc("muted"))
        label_aula_card.configure(text_color=_tc("text"))
        ra_entry.configure(text_color=_tc("text"))
        btn_tema.configure(text="🌙" if _TEMA_ATUAL == "dark" else "☀️", text_color=_tc("text"))
        btn_config.configure(border_color=_tc("card_bor"), text_color=_tc("text"))

        # Propaga text_color="text" para todos os labels dentro dos cards registrados
        def _recolorir(widget):
            try:
                cls = widget.__class__.__name__
                if cls == "CTkLabel":
                    # Só recolore labels sem cor especial (status, verde, vermelho ficam)
                    cur = str(widget.cget("text_color"))
                    if cur in ("#86868b", "#8888aa", "gray", "#555555"):
                        widget.configure(text_color=_tc("muted"))
                    elif cur in ("#f5f5f7", "#1d1d1f", "#0d1f0d", "white", "black"):
                        widget.configure(text_color=_tc("text"))
            except Exception:
                pass
            for child in widget.winfo_children():
                _recolorir(child)
        _recolorir(root)

    def _toggle_tema():
        global _TEMA_ATUAL
        _TEMA_ATUAL = "light" if _TEMA_ATUAL == "dark" else "dark"
        ctk.set_appearance_mode(_TEMA_ATUAL)
        _aplicar_tema_completo()

    # Layout centralizado
    frame_main = ctk.CTkFrame(root, fg_color="transparent")
    frame_main.place(relx=0.5, rely=0.5, anchor="center", relwidth=0.9, relheight=0.95)
    frame_main.columnconfigure(0, weight=1)
    frame_main.rowconfigure(1, weight=1)

    # ── 1. HEADER ──
    hdr = ctk.CTkFrame(frame_main, corner_radius=20, border_width=1)
    hdr.grid(row=0, column=0, sticky="ew", pady=(0, 20))
    _registrar_card(hdr, "hdr", "hdr_bor")

    frame_logo = ctk.CTkFrame(hdr, fg_color="transparent")
    frame_logo.pack(side="left", padx=20, pady=12)

    btn_config = ctk.CTkButton(frame_logo, text="\uE713", font=ctk.CTkFont(family="Segoe MDL2 Assets", size=18), width=40, height=40, corner_radius=20, fg_color="transparent", border_width=1, command=abrir_configuracoes)
    btn_config.pack(side="left", padx=(0,10))
    _registrar_card(btn_config, None, "card_bor")

    try:
        from PIL import Image
        _png_hdr  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sapa_novo_icone_cropped.png")
        _pil_hdr  = Image.open(_png_hdr).convert("RGBA")
        _img_sapo_hdr = ctk.CTkImage(light_image=_pil_hdr, dark_image=_pil_hdr, size=(70, 70))
        ctk.CTkLabel(frame_logo, text="", image=_img_sapo_hdr).pack(side="left", padx=(0,10))
    except Exception:
        ctk.CTkLabel(frame_logo, text="🐸", font=ctk.CTkFont(size=32)).pack(side="left", padx=(0,10))

    frame_txt = ctk.CTkFrame(frame_logo, fg_color="transparent")
    frame_txt.pack(side="left")
    lbl_titulo_hdr = ctk.CTkLabel(frame_txt, text="SAPA", font=ctk.CTkFont(size=22, weight="bold"))
    lbl_titulo_hdr.pack(anchor="w")
    lbl_sub_hdr = ctk.CTkLabel(frame_txt, text="Sistema de Automação de Presença Acadêmica — UNISEPE", font=ctk.CTkFont(size=10))
    lbl_sub_hdr.pack(anchor="w")

    frame_dir = ctk.CTkFrame(hdr, fg_color="transparent")
    frame_dir.pack(side="right", padx=20)

    # Toggle Theme
    btn_tema = ctk.CTkButton(frame_dir, text="🌙" if _TEMA_ATUAL == "dark" else "☀️", font=ctk.CTkFont(size=18), width=40, height=40, corner_radius=20, fg_color="transparent", border_width=1, command=_toggle_tema)
    btn_tema.pack(side="left", padx=10)
    _registrar_card(btn_tema, None, "card_bor")

    info_dir = ctk.CTkFrame(frame_dir, fg_color="transparent")
    info_dir.pack(side="left", padx=10)
    ctk.CTkLabel(info_dir, text="UNISEPE - ADS | F11 - Tela cheia", font=ctk.CTkFont(size=10), text_color=COR_MUTED).pack(anchor="e")
    lbl_wifi = ctk.CTkLabel(info_dir, text="🟢 Online", font=ctk.CTkFont(size=12, weight="bold"), text_color=COR_VERDE)
    lbl_wifi.pack(anchor="e")

    # Scrollable Content se precisar, mas vamos usar frame normal
    content = ctk.CTkFrame(frame_main, fg_color="transparent")
    content.grid(row=1, column=0, sticky="nsew")
    content.columnconfigure(0, weight=1)

    # ── 2. CARD 1: AULA EM ANDAMENTO ──
    card_aula = ctk.CTkFrame(content, corner_radius=24, border_width=1)
    card_aula.pack(fill="x", pady=(0, 12))
    _registrar_card(card_aula, "card", "card_bor")

    c1_inner = ctk.CTkFrame(card_aula, fg_color="transparent")
    c1_inner.pack(fill="both", expand=True, padx=24, pady=14)

    icon_aula = ctk.CTkFrame(c1_inner, width=60, height=60, corner_radius=20, fg_color=("#e6f4ea", "#1b2a20"))
    icon_aula.pack(side="left", padx=(0,20))
    icon_aula.pack_propagate(False)
    ctk.CTkLabel(icon_aula, text="\uE787", font=ctk.CTkFont(family="Segoe MDL2 Assets", size=24), text_color=COR_AZUL).place(relx=0.5, rely=0.5, anchor="center")

    txt_aula = ctk.CTkFrame(c1_inner, fg_color="transparent")
    txt_aula.pack(side="left", fill="x", expand=True)
    ctk.CTkLabel(txt_aula, text="AULA EM ANDAMENTO", font=ctk.CTkFont(size=11, weight="bold"), text_color=COR_MUTED).pack(anchor="w")
    label_aula_card = ctk.CTkLabel(txt_aula, text="Sem aula programada para hoje", font=ctk.CTkFont(size=18, weight="bold"))
    label_aula_card.pack(anchor="w", pady=(2,8))
    _ui_refs["label_aula_card"] = label_aula_card

    # Integração do Combobox e Leitor
    frame_cb = ctk.CTkFrame(txt_aula, fg_color="transparent")
    frame_cb.pack(fill="x")
    combo_aula = ctk.CTkComboBox(frame_cb, command=ao_selecionar_disciplina, font=ctk.CTkFont(size=13), height=38, corner_radius=8)
    combo_aula.set("Selecione a disciplina")
    combo_aula.pack(side="left", expand=True, fill="x", padx=(0,10))
    _ui_refs["combo_aula_ref"] = combo_aula

    ctk.CTkButton(frame_cb, text="🔌 Leitor", command=abrir_painel_leitor, width=100, height=38, corner_radius=8, fg_color=COR_AZUL, font=ctk.CTkFont(size=13)).pack(side="right")

    # ── 3. CARD 2: REGISTRO DE PRESENÇA ──
    card_leitura = ctk.CTkFrame(content, corner_radius=24, border_width=1)
    card_leitura.pack(fill="both", expand=True, pady=(0, 12))
    _registrar_card(card_leitura, "card", "card_bor")

    c2_inner = ctk.CTkFrame(card_leitura, fg_color="transparent")
    c2_inner.pack(fill="both", expand=True, padx=24, pady=12)

    # Header Leitura
    hdr_leit = ctk.CTkFrame(c2_inner, fg_color="transparent")
    hdr_leit.pack(fill="x", pady=(0, 10))
    icon_reg = ctk.CTkFrame(hdr_leit, width=40, height=40, corner_radius=12, fg_color=("#e6f4ea", "#1b2a20"))
    icon_reg.pack(side="left", padx=(0,15))
    icon_reg.pack_propagate(False)
    ctk.CTkLabel(icon_reg, text="\uE77B", font=ctk.CTkFont(family="Segoe MDL2 Assets", size=20), text_color=COR_AZUL).place(relx=0.5, rely=0.5, anchor="center")
    ctk.CTkLabel(hdr_leit, text="REGISTRO DE PRESENÇA", font=ctk.CTkFont(size=11, weight="bold"), text_color=COR_MUTED).pack(side="left")

    # Scan Area Centralizada
    scan_area = ctk.CTkFrame(c2_inner, fg_color="transparent")
    scan_area.pack(pady=4, expand=True)
    frame_msg = ctk.CTkFrame(scan_area, fg_color="transparent")
    frame_msg.pack()
    try:
        from PIL import Image
        _png_scan  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sapa_novo_icone_cropped.png")
        _pil_scan  = Image.open(_png_scan).convert("RGBA")
        _img_sapo_scan = ctk.CTkImage(light_image=_pil_scan, dark_image=_pil_scan, size=(60, 60))
        ctk.CTkLabel(frame_msg, text="", image=_img_sapo_scan).pack(side="left", padx=(0,12))
        ctk.CTkLabel(frame_msg, text="Aponte a carteirinha no leitor", font=ctk.CTkFont(size=22, weight="bold"), text_color="#34a853").pack(side="left")
    except Exception:
        ctk.CTkLabel(frame_msg, text="🐸 Aponte a carteirinha no leitor", font=ctk.CTkFont(size=18, weight="bold"), text_color="#34a853").pack(side="left")

    # Entry Box Estilizada
    ra_entry = ctk.CTkEntry(scan_area, width=380, height=38, font=ctk.CTkFont(size=18), justify="center", corner_radius=8, border_color="#34a853", border_width=2, fg_color="transparent", placeholder_text="RA + Enter")
    ra_entry.pack(pady=6)
    ra_entry.bind("<Return>", _on_entry_ra_return)

    ctk.CTkLabel(scan_area, text="Aproxime a carteirinha do leitor para registrar presença", font=ctk.CTkFont(size=11), text_color=COR_MUTED).pack()

    status_label = ctk.CTkLabel(scan_area, text="", font=ctk.CTkFont(size=13, weight="bold"), text_color=COR_AZUL)
    status_label.pack(pady=(6,0))

    # ── 4. CARD 3: INFO ──
    card_info = ctk.CTkFrame(content, corner_radius=24, border_width=1)
    card_info.pack(fill="x", pady=(0, 12))
    _registrar_card(card_info, "card", "card_bor")

    c3_inner = ctk.CTkFrame(card_info, fg_color="transparent")
    c3_inner.pack(fill="both", expand=True, padx=24, pady=12)

    icon_info = ctk.CTkFrame(c3_inner, width=60, height=60, corner_radius=20, fg_color=("#e6f4ea", "#1b2a20"))
    icon_info.pack(side="left", padx=(0,20))
    icon_info.pack_propagate(False)
    ctk.CTkLabel(icon_info, text="\uE946", font=ctk.CTkFont(family="Segoe MDL2 Assets", size=26), text_color=COR_AZUL).place(relx=0.5, rely=0.5, anchor="center")

    txt_info = ctk.CTkFrame(c3_inner, fg_color="transparent")
    txt_info.pack(side="left", fill="x", expand=True)
    ctk.CTkLabel(txt_info, text="Aponte a carteirinha no leitor", font=ctk.CTkFont(size=18, weight="bold")).pack(anchor="w")
    ctk.CTkLabel(txt_info, text="O sistema registra automaticamente sua presença na aula atual.", font=ctk.CTkFont(size=14), text_color=COR_MUTED).pack(anchor="w")

    # Função para propagar o clique de sub-widgets para o botão pai
    def _vincular_clique_recursivo(widget, handler):
        widget.bind("<Button-1>", lambda e: handler())
        try:
            widget.configure(cursor="hand2")
        except:
            pass
        for child in widget.winfo_children():
            _vincular_clique_recursivo(child, handler)

    # ── 5. ROW DE BOTÕES: ADMIN & PROFESSOR ──
    row_btns = ctk.CTkFrame(content, fg_color="transparent")
    row_btns.pack(fill="x", pady=(0,10))
    row_btns.columnconfigure(0, weight=1)
    row_btns.columnconfigure(1, weight=1)

    # Botão Admin
    btn_admin = ctk.CTkButton(
        row_btns, text="", corner_radius=24, border_width=1, height=100,
        command=lambda: pedir_login("admin"),
        fg_color=("#f0f6f2", "#142819"),
        border_color=("#ffffff", "#1a1a1a"),
        hover_color=("#e5e5e5", "#2a2a3d")
    )
    btn_admin.grid(row=0, column=0, sticky="ew", padx=(0,10))

    in_adm = ctk.CTkFrame(btn_admin, fg_color="transparent")
    in_adm.place(relx=0, rely=0, relwidth=1, relheight=1)

    ic_adm = ctk.CTkFrame(in_adm, width=60, height=60, corner_radius=20, fg_color=("#e6f4ea", "#1b2a20"))
    ic_adm.place(x=20, rely=0.5, anchor="w")
    ic_adm.pack_propagate(False)
    ctk.CTkLabel(ic_adm, text="\uEA18", font=ctk.CTkFont(family="Segoe MDL2 Assets", size=26), text_color=COR_AZUL).place(relx=0.5, rely=0.5, anchor="center")

    ctk.CTkLabel(in_adm, text="Administração", font=ctk.CTkFont(size=16, weight="bold"), text_color=COR_AZUL).place(x=95, rely=0.35, anchor="w")
    ctk.CTkLabel(in_adm, text="Configurações do sistema", font=ctk.CTkFont(size=12), text_color=COR_MUTED).place(x=95, rely=0.65, anchor="w")
    ctk.CTkLabel(in_adm, text=">", font=ctk.CTkFont(size=20, weight="bold"), text_color=COR_AZUL).place(relx=0.92, rely=0.5, anchor="e")

    _vincular_clique_recursivo(in_adm, lambda: pedir_login("admin"))

    # Botão Professor
    btn_prof = ctk.CTkButton(
        row_btns, text="", corner_radius=24, border_width=1, height=100,
        command=lambda: pedir_login("professor"),
        fg_color=("#f0f6f2", "#142819"),
        border_color=("#ffffff", "#1a1a1a"),
        hover_color=("#e5e5e5", "#2a2a3d")
    )
    btn_prof.grid(row=0, column=1, sticky="ew", padx=(10,0))

    in_pro = ctk.CTkFrame(btn_prof, fg_color="transparent")
    in_pro.place(relx=0, rely=0, relwidth=1, relheight=1)

    ic_pro = ctk.CTkFrame(in_pro, width=60, height=60, corner_radius=20, fg_color=("#e6f4ea", "#1b2a20"))
    ic_pro.place(x=20, rely=0.5, anchor="w")
    ic_pro.pack_propagate(False)
    ctk.CTkLabel(ic_pro, text="\uE716", font=ctk.CTkFont(family="Segoe MDL2 Assets", size=26), text_color=COR_AZUL).place(relx=0.5, rely=0.5, anchor="center")

    ctk.CTkLabel(in_pro, text="Painel do Professor", font=ctk.CTkFont(size=16, weight="bold"), text_color=COR_AZUL).place(x=95, rely=0.35, anchor="w")
    ctk.CTkLabel(in_pro, text="Gerencie suas turmas e aulas", font=ctk.CTkFont(size=12), text_color=COR_MUTED).place(x=95, rely=0.65, anchor="w")
    ctk.CTkLabel(in_pro, text=">", font=ctk.CTkFont(size=20, weight="bold"), text_color=COR_AZUL).place(relx=0.92, rely=0.5, anchor="e")

    _vincular_clique_recursivo(in_pro, lambda: pedir_login("professor"))

    _aplicar_tema_completo()

    return root


# ──────────────────────────────────────────────────────────
# MIGRAÇÃO DO BANCO (registro único)
# ──────────────────────────────────────────────────────────
def migrar_logs_para_registro_unico():
    """
    Adiciona as colunas Hora_Entrada e Hora_Saida na tabela LOGS
    e preenche Hora_Entrada a partir da coluna Hora existente.
    """
    with conectar() as conn:
        for col in ["Hora_Entrada","Hora_Saida","Justificativa"]:
            try: conn.execute(f"ALTER TABLE LOGS ADD COLUMN {col} TEXT")
            except sqlite3.OperationalError: pass
        # Copia Hora → Hora_Entrada onde Tipo='ENTRADA' e Hora_Entrada é nulo
        conn.execute("""
            UPDATE LOGS SET Hora_Entrada = Hora
            WHERE Tipo = 'ENTRADA' AND (Hora_Entrada IS NULL OR Hora_Entrada = '')
        """)
        conn.execute("""
            UPDATE LOGS SET Hora_Saida = Hora
            WHERE Tipo = 'SAIDA' AND (Hora_Saida IS NULL OR Hora_Saida = '')
        """)
        conn.commit()

def migrar_disciplinas_para_novos_campos():
    with conectar() as conn:
        for col in ["Data_Inicio", "Data_Fim", "Dia_Semana"]:
            try: conn.execute(f"ALTER TABLE DISCIPLINAS ADD COLUMN {col} TEXT")
            except sqlite3.OperationalError: pass
        conn.commit()

# ──────────────────────────────────────────────────────────
# ARRANQUE
# ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    inicializar_fila_local()
    garantir_schema()
    migrar_logs_para_registro_unico()
    migrar_disciplinas_para_novos_campos()
    iniciar_thread_sync(intervalo_segundos=15)
    iniciar_robo_faltas()

    root = construir_ui()
    atualizar_menu_principal()    # popula o combo com as disciplinas do banco
    atualizar_disciplina_automatica()
    iniciar_monitor_wifi()

    root.mainloop()
